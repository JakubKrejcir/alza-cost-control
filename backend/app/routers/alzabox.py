"""
AlzaBox Router - Import dat a BI API (ASYNC verze)
Verze: 3.13.0 - Fixed carrier filtering to use AlzaBoxDelivery.carrierId directly
Updated: 2025-12-09
"""
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, delete, text
from datetime import datetime, timedelta, date
from typing import Optional, List
from decimal import Decimal
import openpyxl
import io
import re
import logging

from app.database import get_db
from app.models import AlzaBox, AlzaBoxAssignment, AlzaBoxDelivery, Carrier
from app.carrier_matching import build_carrier_lookup, find_carrier_id

router = APIRouter(tags=["alzabox"])
logger = logging.getLogger(__name__)


# =============================================================================
# IMPORT ENDPOINTS
# =============================================================================

@router.post("/import/locations")
async def import_alzabox_locations(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Import AlzaBox umístění z XLSX souboru.
    
    Očekávaná struktura (sheet LL_PS):
    - Col 1: Výdejní místo (id) → alza_id
    - Col 2: Kód dopravce → code (AB130)
    - Col 4: Skupina výdejních míst → route_group
    - Col 5: Odesílatel → source_warehouse
    - Col 6: Dopravce → carrier_name
    - Col 7: Název combo → name
    - Col 8: Stát → country
    - Col 9: Město → city
    - Col 10: GPS Y → gps_lat
    - Col 11: GPS X → gps_lon
    - Col 12: Kraj → region
    - Col 13: První spuštění → first_launch
    """
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        # Flexibilní detekce sheetu
        sheet = None
        for name in ['LL_PS', 'Sheet1', 'Data', 'Boxy', 'AlzaBoxy', 'List1']:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        
        if not sheet:
            sheet = wb.active
        
        if not sheet:
            raise HTTPException(status_code=400, detail=f"Žádný sheet nenalezen. Dostupné: {wb.sheetnames}")
        
        logger.info(f"Importing locations from sheet: {sheet.title}, rows: {sheet.max_row}")
        
        # Načti existující dopravce - lookup podle name I alias
        result = await db.execute(select(Carrier))
        carriers_list = result.scalars().all()
        carrier_lookup = build_carrier_lookup(carriers_list)
        logger.info(f"Found {len(carriers_list)} carriers in DB")
        
        # Načti existující boxy
        result = await db.execute(select(AlzaBox))
        existing_boxes = {b.code: b for b in result.scalars().all()}
        logger.info(f"Found {len(existing_boxes)} existing boxes")
        
        created = 0
        updated = 0
        skipped = 0
        
        for row in range(2, sheet.max_row + 1):
            alza_id = sheet.cell(row=row, column=1).value
            code = sheet.cell(row=row, column=2).value
            route_group = sheet.cell(row=row, column=4).value
            source_warehouse = sheet.cell(row=row, column=5).value
            carrier_name = sheet.cell(row=row, column=6).value
            name = sheet.cell(row=row, column=7).value
            country = sheet.cell(row=row, column=8).value
            city = sheet.cell(row=row, column=9).value
            gps_lat = sheet.cell(row=row, column=10).value
            gps_lon = sheet.cell(row=row, column=11).value
            region = sheet.cell(row=row, column=12).value
            first_launch = sheet.cell(row=row, column=13).value
            
            if not code:
                skipped += 1
                continue
            
            code = str(code).strip()
            name = str(name).strip() if name else code
            
            if code in existing_boxes:
                box = existing_boxes[code]
                box.name = name
                box.country = country or 'CZ'
                box.city = city
                box.region = region
                box.gps_lat = Decimal(str(gps_lat)) if gps_lat else None
                box.gps_lon = Decimal(str(gps_lon)) if gps_lon else None
                box.source_warehouse = source_warehouse
                box.updated_at = datetime.utcnow()
                updated += 1
            else:
                box = AlzaBox(
                    code=code,
                    alza_id=alza_id,
                    name=name,
                    country=country or 'CZ',
                    city=city,
                    region=region,
                    gps_lat=Decimal(str(gps_lat)) if gps_lat else None,
                    gps_lon=Decimal(str(gps_lon)) if gps_lon else None,
                    source_warehouse=source_warehouse,
                    first_launch=first_launch if isinstance(first_launch, datetime) else None,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow()
                )
                db.add(box)
                await db.flush()
                existing_boxes[code] = box
                created += 1
            
            # Přiřazení k dopravci (hledá podle name i alias)
            carrier_id = find_carrier_id(carrier_name, carrier_lookup) if carrier_name else None
            
            if carrier_id or route_group:
                assignment = AlzaBoxAssignment(
                    box_id=box.id,
                    carrier_id=carrier_id,
                    route_group=str(route_group) if route_group else None,
                    depot_name=None,
                    valid_from=datetime.utcnow(),
                    created_at=datetime.utcnow()
                )
                db.add(assignment)
        
        await db.commit()
        wb.close()
        
        logger.info(f"Import complete: created={created}, updated={updated}, skipped={skipped}")
        
        return {
            "success": True,
            "imported": created + updated,
            "boxes_created": created,
            "boxes_updated": updated,
            "skipped": skipped
        }
        
    except Exception as e:
        await db.rollback()
        logger.error(f"Error importing locations: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import/deliveries")
async def import_alzabox_deliveries(
    file: UploadFile = File(...),
    delivery_type: str = Query("DPO", description="Typ závozu: DPO, SD, THIRD"),
    db: AsyncSession = Depends(get_db)
):
    """
    Import dojezdových časů z XLSX souboru.
    
    Očekávaná struktura:
    - Sheet 'Plan' a 'Actual'
    - Row 2: datumy (od col 2)
    - Row 3+: data
      - Col 1: "Název trasy" (hlavička skupiny) NEBO "09:00 | Název boxu -- AB1234"
      - Col 2+: časy dojezdů (datetime)
    """
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        logger.info(f"Available sheets: {wb.sheetnames}")
        
        # Flexibilní detekce sheetů
        plan_sheet = None
        actual_sheet = None
        
        for name in ['Plan', 'Plán', 'plan', 'PLAN']:
            if name in wb.sheetnames:
                plan_sheet = wb[name]
                break
        
        for name in ['Actual', 'Skutecnost', 'Skutečnost', 'actual', 'ACTUAL']:
            if name in wb.sheetnames:
                actual_sheet = wb[name]
                break
        
        if not plan_sheet or not actual_sheet:
            available = ', '.join(wb.sheetnames)
            raise HTTPException(
                status_code=400, 
                detail=f"Sheety 'Plan' a 'Actual' nenalezeny. Dostupné: {available}"
            )
        
        logger.info(f"Using sheets: Plan={plan_sheet.title}, Actual={actual_sheet.title}")
        
        # Načti datumy z ROW 2 (od col 2)
        dates = []
        for col in range(2, plan_sheet.max_column + 1):
            date_val = plan_sheet.cell(row=2, column=col).value
            if isinstance(date_val, datetime):
                dates.append((col, date_val.date()))
        
        if not dates:
            raise HTTPException(status_code=400, detail="Žádné datumy nenalezeny v řádku 2")
        
        logger.info(f"Found {len(dates)} dates: {dates[0][1]} to {dates[-1][1]}")
        
        # Smaž existující záznamy pro období
        min_date = min(d for _, d in dates)
        max_date = max(d for _, d in dates)
        
        delete_stmt = delete(AlzaBoxDelivery).where(
            and_(
                AlzaBoxDelivery.delivery_type == delivery_type,
                AlzaBoxDelivery.delivery_date >= datetime.combine(min_date, datetime.min.time()),
                AlzaBoxDelivery.delivery_date <= datetime.combine(max_date, datetime.max.time())
            )
        )
        result = await db.execute(delete_stmt)
        await db.flush()
        logger.info(f"Deleted {result.rowcount} existing records")
        
        # Načti boxy z DB
        result = await db.execute(select(AlzaBox))
        boxes = {b.code: b.id for b in result.scalars().all()}
        logger.info(f"Loaded {len(boxes)} boxes from DB")
        
        # Parsuj Plan sheet - extrahuj plánované časy a kódy boxů
        # Formát: "09:00 | Brno - Bystrc (OC Max) -- AB1688"
        plan_data = {}
        current_route = None
        
        for row in range(3, plan_sheet.max_row + 1):
            col1 = plan_sheet.cell(row=row, column=1).value
            
            if not col1:
                continue
            
            col1 = str(col1).strip()
            
            # Detekuj hlavičku trasy (nemá | ani --)
            if '|' not in col1 and '--' not in col1:
                current_route = col1
                continue
            
            # Parsuj řádek boxu: "09:00 | Název -- AB1234"
            # Regex pro extrakci času a kódu
            match = re.match(r'^(\d{1,2}:\d{2})\s*\|\s*.*--\s*(AB\d+)', col1)
            if match:
                planned_time = match.group(1)
                box_code = match.group(2)
                plan_data[box_code] = {
                    'route_name': current_route,
                    'planned_time': planned_time,
                    'row': row
                }
        
        logger.info(f"Parsed {len(plan_data)} boxes from Plan sheet")
        
        # Parsuj Actual sheet - stejná struktura
        actual_data = {}
        
        for row in range(3, actual_sheet.max_row + 1):
            col1 = actual_sheet.cell(row=row, column=1).value
            
            if not col1:
                continue
            
            col1 = str(col1).strip()
            
            # Parsuj řádek boxu
            match = re.match(r'^(\d{1,2}:\d{2})\s*\|\s*.*--\s*(AB\d+)', col1)
            if match:
                box_code = match.group(2)
                actual_data[box_code] = {'row': row}
                
                # Načti časy pro všechny datumy
                for col, dt in dates:
                    val = actual_sheet.cell(row=row, column=col).value
                    if val and isinstance(val, datetime):
                        if 'times' not in actual_data[box_code]:
                            actual_data[box_code]['times'] = {}
                        actual_data[box_code]['times'][col] = val
        
        logger.info(f"Parsed {len(actual_data)} boxes from Actual sheet")
        
        # Vytvoř delivery záznamy
        created = 0
        skipped_no_box = 0
        skipped_no_actual = 0
        
        for box_code, plan_info in plan_data.items():
            if box_code not in boxes:
                skipped_no_box += 1
                continue
            
            box_id = boxes[box_code]
            route_name = plan_info['route_name']
            planned_time = plan_info['planned_time']
            
            if box_code not in actual_data or 'times' not in actual_data[box_code]:
                skipped_no_actual += 1
                continue
            
            for col, dt in dates:
                if col not in actual_data[box_code]['times']:
                    continue
                
                actual_time = actual_data[box_code]['times'][col]
                
                # Vypočítej zpoždění
                delay_minutes = None
                on_time = None
                if planned_time:
                    try:
                        parts = planned_time.split(':')
                        plan_min = int(parts[0]) * 60 + int(parts[1])
                        actual_min = actual_time.hour * 60 + actual_time.minute
                        delay_minutes = actual_min - plan_min
                        on_time = delay_minutes <= 0
                    except:
                        pass
                
                delivery = AlzaBoxDelivery(
                    box_id=box_id,
                    delivery_date=datetime.combine(dt, actual_time.time()),
                    delivery_type=delivery_type,
                    route_name=route_name,
                    planned_time=planned_time,
                    actual_time=actual_time,
                    delay_minutes=delay_minutes,
                    on_time=on_time,
                    created_at=datetime.utcnow()
                )
                db.add(delivery)
                created += 1
        
        await db.commit()
        wb.close()
        
        logger.info(f"Import complete: created={created}, skipped_no_box={skipped_no_box}, skipped_no_actual={skipped_no_actual}")
        
        return {
            "success": True,
            "imported": created,
            "skipped_no_box": skipped_no_box,
            "skipped_no_actual": skipped_no_actual,
            "dates_processed": len(dates)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error importing deliveries: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# DELETE ENDPOINTS
# =============================================================================

@router.delete("/data/locations")
async def delete_all_locations(db: AsyncSession = Depends(get_db)):
    """Smaže všechny AlzaBoxy"""
    try:
        await db.execute(delete(AlzaBoxDelivery))
        await db.execute(delete(AlzaBoxAssignment))
        result = await db.execute(delete(AlzaBox))
        await db.commit()
        return {"success": True, "deleted": result.rowcount}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/data/deliveries")
async def delete_deliveries(
    delivery_type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Smaže dojezdy"""
    try:
        if delivery_type:
            result = await db.execute(
                delete(AlzaBoxDelivery).where(AlzaBoxDelivery.delivery_type == delivery_type)
            )
        else:
            result = await db.execute(delete(AlzaBoxDelivery))
        await db.commit()
        return {"success": True, "deleted": result.rowcount}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# STATS ENDPOINTS
# =============================================================================

@router.get("/stats/summary")
async def get_summary(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    delivery_type: Optional[str] = Query(None),
    carrier_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Celkový přehled - počet boxů, dojezdů, včasnost"""
    try:
        # Počet boxů
        box_result = await db.execute(text('SELECT COUNT(*) FROM "AlzaBox"'))
        total_boxes = box_result.scalar() or 0
        
        # Statistiky dojezdů
        sql = """
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN "onTime" = true THEN 1 ELSE 0 END) as on_time,
            AVG("delayMinutes") as avg_delay
        FROM "AlzaBoxDelivery"
        WHERE 1=1
        """
        params = {}
        
        if start_date:
            sql += ' AND "deliveryDate" >= :start_date'
            params['start_date'] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            sql += ' AND "deliveryDate" <= :end_date'
            params['end_date'] = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        if delivery_type:
            sql += ' AND "deliveryType" = :delivery_type'
            params['delivery_type'] = delivery_type
        if carrier_id:
            sql += ' AND "carrierId" = :carrier_id'
            params['carrier_id'] = carrier_id
        
        result = await db.execute(text(sql), params)
        row = result.fetchone()
        
        total_deliveries = row[0] or 0
        on_time = row[1] or 0
        on_time_rate = round(on_time / total_deliveries * 100, 1) if total_deliveries > 0 else 0
        
        return {
            "total_boxes": total_boxes,
            "total_deliveries": total_deliveries,
            "on_time_rate": on_time_rate,
            # Zachováme i staré názvy pro kompatibilitu
            "totalDeliveries": total_deliveries,
            "onTimeDeliveries": on_time,
            "onTimePct": on_time_rate
        }
    except Exception as e:
        logger.error(f"Error in summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats/by-carrier")
async def get_by_carrier(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Statistiky per dopravce"""
    try:
        sql = """
        SELECT 
            c.name,
            c.id,
            COUNT(d.id) as total,
            SUM(CASE WHEN d."onTime" = true THEN 1 ELSE 0 END) as on_time,
            AVG(CASE WHEN d."onTime" = false THEN d."delayMinutes" ELSE NULL END) as avg_delay
        FROM "AlzaBoxDelivery" d
        LEFT JOIN "Carrier" c ON d."carrierId" = c.id
        WHERE 1=1
        """
        params = {}
        
        if start_date:
            sql += ' AND d."deliveryDate" >= :start_date'
            params['start_date'] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            sql += ' AND d."deliveryDate" <= :end_date'
            params['end_date'] = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        
        sql += ' GROUP BY c.id, c.name ORDER BY total DESC'
        
        result = await db.execute(text(sql), params)
        rows = result.fetchall()
        
        return [
            {
                "carrierName": row[0] or "Nepřiřazeno",
                "carrierId": row[1],
                "totalDeliveries": row[2],
                "onTimeDeliveries": row[3] or 0,
                "onTimePct": round((row[3] or 0) / row[2] * 100, 1) if row[2] > 0 else 0,
                "avgDelayMinutes": round(float(row[4]), 1) if row[4] else None
            }
            for row in rows
        ]
    except Exception as e:
        logger.error(f"Error in by-carrier: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats/by-route")
async def get_by_route(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    carrier_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Statistiky per trasa"""
    try:
        sql = """
        SELECT 
            d."routeName",
            COUNT(*) as total,
            SUM(CASE WHEN d."onTime" = true THEN 1 ELSE 0 END) as on_time,
            AVG(CASE WHEN d."onTime" = false THEN d."delayMinutes" ELSE NULL END) as avg_delay
        FROM "AlzaBoxDelivery" d
        WHERE d."routeName" IS NOT NULL
        """
        params = {}
        
        if start_date:
            sql += ' AND d."deliveryDate" >= :start_date'
            params['start_date'] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            sql += ' AND d."deliveryDate" <= :end_date'
            params['end_date'] = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        if carrier_id:
            sql += ' AND d."carrierId" = :carrier_id'
            params['carrier_id'] = carrier_id
        
        sql += ' GROUP BY d."routeName" ORDER BY d."routeName"'
        
        result = await db.execute(text(sql), params)
        rows = result.fetchall()
        
        return [
            {
                "routeName": row[0],
                "totalDeliveries": row[1],
                "onTimeDeliveries": row[2] or 0,
                "onTimePct": round((row[2] or 0) / row[1] * 100, 1) if row[1] > 0 else 0,
                "avgDelayMinutes": round(float(row[3]), 1) if row[3] else None
            }
            for row in rows
        ]
    except Exception as e:
        logger.error(f"Error in by-route: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats/by-day")
async def get_by_day(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    carrier_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Statistiky per den"""
    try:
        sql = """
        SELECT 
            DATE(d."deliveryDate") as date,
            COUNT(*) as total,
            SUM(CASE WHEN d."onTime" = true THEN 1 ELSE 0 END) as on_time
        FROM "AlzaBoxDelivery" d
        WHERE 1=1
        """
        params = {}
        
        if start_date:
            sql += ' AND d."deliveryDate" >= :start_date'
            params['start_date'] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            sql += ' AND d."deliveryDate" <= :end_date'
            params['end_date'] = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        if carrier_id:
            sql += ' AND d."carrierId" = :carrier_id'
            params['carrier_id'] = carrier_id
        
        sql += ' GROUP BY DATE(d."deliveryDate") ORDER BY date'
        
        result = await db.execute(text(sql), params)
        rows = result.fetchall()
        
        return [
            {
                "date": str(row[0]),
                "totalDeliveries": row[1],
                "onTimeDeliveries": row[2] or 0,
                "onTimePct": round((row[2] or 0) / row[1] * 100, 1) if row[1] > 0 else 0
            }
            for row in rows
        ]
    except Exception as e:
        logger.error(f"Error in by-day: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats/by-box")
async def get_by_box(
    route_name: Optional[str] = Query(None),
    carrier_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Statistiky per box"""
    try:
        sql = """
        SELECT 
            b.id, b.code, b.name, b.city,
            COUNT(d.id) as total,
            SUM(CASE WHEN d."onTime" = true THEN 1 ELSE 0 END) as on_time
        FROM "AlzaBox" b
        JOIN "AlzaBoxDelivery" d ON b.id = d."boxId"
        WHERE 1=1
        """
        params = {}
        
        if route_name:
            sql += ' AND d."routeName" = :route_name'
            params['route_name'] = route_name
        if carrier_id:
            sql += ' AND d."carrierId" = :carrier_id'
            params['carrier_id'] = carrier_id
        
        sql += ' GROUP BY b.id, b.code, b.name, b.city ORDER BY b.code'
        
        result = await db.execute(text(sql), params)
        rows = result.fetchall()
        
        return [
            {
                "boxId": row[0],
                "boxCode": row[1],
                "boxName": row[2],
                "city": row[3],
                "totalDeliveries": row[4],
                "onTimeDeliveries": row[5] or 0,
                "onTimePct": round((row[5] or 0) / row[4] * 100, 1) if row[4] > 0 else 0
            }
            for row in rows
        ]
    except Exception as e:
        logger.error(f"Error in by-box: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/box/{box_id}/detail")
async def get_box_detail(box_id: int, db: AsyncSession = Depends(get_db)):
    """Detail boxu"""
    try:
        # Box info
        result = await db.execute(text(
            'SELECT id, code, name, city, region FROM "AlzaBox" WHERE id = :id'
        ), {"id": box_id})
        box = result.fetchone()
        
        if not box:
            raise HTTPException(status_code=404, detail="Box nenalezen")
        
        # Dopravce
        result = await db.execute(text("""
            SELECT c.id, c.name FROM "Carrier" c
            JOIN "AlzaBoxAssignment" a ON c.id = a."carrierId"
            WHERE a."boxId" = :id AND a."validTo" IS NULL LIMIT 1
        """), {"id": box_id})
        carrier = result.fetchone()
        
        # Dojezdy
        result = await db.execute(text("""
            SELECT "deliveryDate", "routeName", "plannedTime", "actualTime", "delayMinutes", "onTime"
            FROM "AlzaBoxDelivery" WHERE "boxId" = :id
            ORDER BY "deliveryDate" DESC LIMIT 50
        """), {"id": box_id})
        deliveries = result.fetchall()
        
        total = len(deliveries)
        on_time = sum(1 for d in deliveries if d[5])
        
        return {
            "box": {"id": box[0], "code": box[1], "name": box[2], "city": box[3], "region": box[4]},
            "carrier": {"id": carrier[0], "name": carrier[1]} if carrier else None,
            "stats": {
                "totalDeliveries": total,
                "onTimeDeliveries": on_time,
                "onTimePct": round(on_time / total * 100, 1) if total > 0 else 0
            },
            "history": [
                {
                    "date": str(d[0].date()) if d[0] else None,
                    "routeName": d[1],
                    "plannedTime": d[2],
                    "actualTime": str(d[3].time()) if d[3] else None,
                    "delayMinutes": d[4],
                    "onTime": d[5]
                }
                for d in deliveries
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in box detail: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/boxes")
async def get_boxes(
    limit: int = Query(100, le=5000),
    offset: int = 0,
    db: AsyncSession = Depends(get_db)
):
    """Seznam boxů"""
    try:
        result = await db.execute(text('SELECT COUNT(*) FROM "AlzaBox"'))
        total = result.scalar()
        
        result = await db.execute(text(
            'SELECT id, code, name, country, city, region FROM "AlzaBox" LIMIT :limit OFFSET :offset'
        ), {"limit": limit, "offset": offset})
        rows = result.fetchall()
        
        return {
            "total": total,
            "boxes": [
                {"id": r[0], "code": r[1], "name": r[2], "country": r[3], "city": r[4], "region": r[5]}
                for r in rows
            ]
        }
    except Exception as e:
        logger.error(f"Error in boxes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/routes")
async def get_routes(carrier_id: Optional[int] = Query(None), db: AsyncSession = Depends(get_db)):
    """Seznam tras"""
    try:
        sql = 'SELECT DISTINCT "routeName" FROM "AlzaBoxDelivery" WHERE "routeName" IS NOT NULL'
        params = {}
        
        if carrier_id:
            sql = """
            SELECT DISTINCT "routeName" FROM "AlzaBoxDelivery"
            WHERE "routeName" IS NOT NULL AND "carrierId" = :carrier_id
            """
            params['carrier_id'] = carrier_id
        
        sql += ' ORDER BY "routeName"'
        
        result = await db.execute(text(sql), params)
        rows = result.fetchall()
        
        return [{"routeName": r[0]} for r in rows]
    except Exception as e:
        logger.error(f"Error in routes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/carriers")
async def get_carriers(db: AsyncSession = Depends(get_db)):
    """Seznam dopravců s AlzaBox přiřazeními"""
    try:
        result = await db.execute(text("""
            SELECT DISTINCT c.id, c.name FROM "Carrier" c
            JOIN "AlzaBoxAssignment" a ON c.id = a."carrierId"
            WHERE a."validTo" IS NULL ORDER BY c.name
        """))
        rows = result.fetchall()
        return [{"id": r[0], "name": r[1]} for r in rows]
    except Exception as e:
        logger.error(f"Error in carriers: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/countries")
async def get_countries(db: AsyncSession = Depends(get_db)):
    """Seznam zemí"""
    try:
        result = await db.execute(text(
            'SELECT country, COUNT(*) FROM "AlzaBox" GROUP BY country ORDER BY COUNT(*) DESC'
        ))
        rows = result.fetchall()
        return [{"country": r[0], "boxCount": r[1]} for r in rows]
    except Exception as e:
        logger.error(f"Error in countries: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/diagnostics/carrier-mapping")
async def get_diagnostics(db: AsyncSession = Depends(get_db)):
    """Diagnostika mapování"""
    try:
        result = await db.execute(text("""
            SELECT c.name, COUNT(a.id), COUNT(DISTINCT a."boxId")
            FROM "AlzaBoxAssignment" a
            LEFT JOIN "Carrier" c ON a."carrierId" = c.id
            WHERE a."validTo" IS NULL
            GROUP BY c.name ORDER BY COUNT(a.id) DESC
        """))
        rows = result.fetchall()
        
        return {
            "assignments": [
                {"carrierName": r[0] or "Nepřiřazeno", "count": r[1], "uniqueBoxes": r[2]}
                for r in rows
            ]
        }
    except Exception as e:
        logger.error(f"Error in diagnostics: {e}")
        raise HTTPException(status_code=500, detail=str(e))
