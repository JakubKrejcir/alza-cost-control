"""
Plans API Router - for comparing planned routes with actual proofs
"""
from typing import List, Optional
from datetime import datetime, date, time, timedelta
from decimal import Decimal
from io import BytesIO
import re
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload, Mapped, mapped_column, relationship
from sqlalchemy import String, Integer, Boolean, DateTime, Date, ForeignKey, Text, Numeric
import openpyxl

from app.database import get_db, Base
from app.models import Carrier, Proof

router = APIRouter()


# =============================================================================
# MODELS
# =============================================================================

class Plan(Base):
    """Planning document - represents a route plan for a date range"""
    __tablename__ = "Plan"

    id: Mapped[int] = mapped_column(primary_key=True)
    carrier_id: Mapped[int] = mapped_column("carrierId", ForeignKey("Carrier.id", ondelete="CASCADE"))
    name: Mapped[str] = mapped_column(String(255))
    
    # Date range for which this plan is valid
    valid_from: Mapped[date] = mapped_column("validFrom", Date)
    valid_to: Mapped[date] = mapped_column("validTo", Date)
    
    period: Mapped[Optional[str]] = mapped_column(String(20))  # MM/YYYY for grouping
    file_name: Mapped[Optional[str]] = mapped_column("fileName", String(255))
    
    # Per-day metrics
    routes_per_day: Mapped[int] = mapped_column("routesPerDay", Integer, default=0)
    linehauls_per_batch: Mapped[int] = mapped_column("linehaulsPerBatch", Integer, default=2)
    distance_per_day_km: Mapped[Optional[Decimal]] = mapped_column("distancePerDayKm", Numeric(10, 2))
    stops_per_day: Mapped[Optional[int]] = mapped_column("stopsPerDay", Integer)
    
    routes_dr_per_day: Mapped[int] = mapped_column("routesDrPerDay", Integer, default=0)
    routes_lh_per_day: Mapped[int] = mapped_column("routesLhPerDay", Integer, default=0)
    
    # Calculated totals for the entire period
    working_days: Mapped[int] = mapped_column("workingDays", Integer, default=1)
    total_routes: Mapped[int] = mapped_column("totalRoutes", Integer, default=0)
    total_distance_km: Mapped[Optional[Decimal]] = mapped_column("totalDistanceKm", Numeric(10, 2))
    
    status: Mapped[str] = mapped_column(String(50), default="active")
    notes: Mapped[Optional[str]] = mapped_column(Text)
    
    created_at: Mapped[datetime] = mapped_column("createdAt", DateTime, default=datetime.utcnow)
    updated_at: Mapped[datetime] = mapped_column("updatedAt", DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    carrier: Mapped["Carrier"] = relationship("Carrier")
    routes: Mapped[List["PlanRoute"]] = relationship("PlanRoute", back_populates="plan", cascade="all, delete-orphan")


class PlanRoute(Base):
    """Individual route in a plan (template for a day)"""
    __tablename__ = "PlanRoute"

    id: Mapped[int] = mapped_column(primary_key=True)
    plan_id: Mapped[int] = mapped_column("planId", ForeignKey("Plan.id", ondelete="CASCADE"))
    
    vehicle_id: Mapped[str] = mapped_column("vehicleId", String(100))
    route_code: Mapped[Optional[str]] = mapped_column("routeCode", String(20))
    route_type: Mapped[str] = mapped_column("routeType", String(50))
    start_location: Mapped[Optional[str]] = mapped_column("startLocation", String(255))
    distance_km: Mapped[Decimal] = mapped_column("distanceKm", Numeric(10, 2), default=0)
    work_time_minutes: Mapped[Optional[int]] = mapped_column("workTimeMinutes", Integer)
    stops_count: Mapped[Optional[int]] = mapped_column("stopsCount", Integer)

    plan: Mapped["Plan"] = relationship("Plan", back_populates="routes")


class PlanComparison(Base):
    """Comparison of aggregated plans with a proof"""
    __tablename__ = "PlanComparison"

    id: Mapped[int] = mapped_column(primary_key=True)
    proof_id: Mapped[int] = mapped_column("proofId", ForeignKey("Proof.id", ondelete="CASCADE"))
    
    # Which plans were included (comma-separated IDs)
    plan_ids: Mapped[str] = mapped_column("planIds", Text)
    
    # Aggregated plan totals
    plans_count: Mapped[int] = mapped_column("plansCount", Integer, default=0)
    total_working_days: Mapped[int] = mapped_column("totalWorkingDays", Integer, default=0)
    routes_planned: Mapped[int] = mapped_column("routesPlanned", Integer, default=0)
    routes_actual: Mapped[int] = mapped_column("routesActual", Integer, default=0)
    routes_difference: Mapped[int] = mapped_column("routesDifference", Integer, default=0)
    
    distance_planned_km: Mapped[Optional[Decimal]] = mapped_column("distancePlannedKm", Numeric(12, 2))
    
    cost_actual: Mapped[Optional[Decimal]] = mapped_column("costActual", Numeric(12, 2))
    
    extra_routes: Mapped[int] = mapped_column("extraRoutes", Integer, default=0)
    missing_routes: Mapped[int] = mapped_column("missingRoutes", Integer, default=0)
    combined_routes: Mapped[int] = mapped_column("combinedRoutes", Integer, default=0)
    
    differences_json: Mapped[Optional[str]] = mapped_column("differencesJson", Text)
    status: Mapped[str] = mapped_column(String(50), default="completed")
    notes: Mapped[Optional[str]] = mapped_column(Text)
    
    created_at: Mapped[datetime] = mapped_column("createdAt", DateTime, default=datetime.utcnow)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def count_working_days(start: date, end: date) -> int:
    """Count working days (Mon-Fri) between two dates inclusive"""
    count = 0
    current = start
    while current <= end:
        if current.weekday() < 5:  # Monday = 0, Friday = 4
            count += 1
        current += timedelta(days=1)
    return max(count, 1)


def parse_plan_from_xlsx(file_content: bytes, plan_date: date) -> dict:
    """Parse planning data from XLSX file
    
    LH-LH = 2 linehaul TRUCKS for ALL routes in the batch (not per route!)
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    
    result = {
        'routes': [],
        'total_routes': 0,
        'linehauls_per_batch': 0,
        'total_distance_km': Decimal('0'),
        'total_stops': 0,
        'routes_dr': 0,
        'routes_lh': 0,
    }
    
    # Find Routes sheet
    routes_sheet = None
    for name in ['Routes', 'Trasy', 'routes']:
        if name in wb.sheetnames:
            routes_sheet = wb[name]
            break
    
    if not routes_sheet:
        routes_sheet = wb.active
    
    # Find header row
    header_row = None
    headers = {}
    for row_idx in range(1, min(10, routes_sheet.max_row + 1)):
        row = [routes_sheet.cell(row=row_idx, column=c).value for c in range(1, routes_sheet.max_column + 1)]
        if any('vozidla' in str(cell).lower() for cell in row if cell) or \
           any('vehicle' in str(cell).lower() for cell in row if cell):
            header_row = row_idx
            for col_idx, cell in enumerate(row):
                if cell:
                    headers[str(cell).lower().strip()] = col_idx
            break
    
    if not header_row:
        raise ValueError("Could not find header row in Routes sheet")
    
    # Map column names
    col_map = {
        'vehicle_id': next((headers.get(k) for k in ['identifikátor vozidla', 'vehicle', 'vozidlo', 'trasa'] if k in headers), None),
        'carrier': next((headers.get(k) for k in ['dopravce', 'carrier'] if k in headers), None),
        'stops': next((headers.get(k) for k in ['náklad 1', 'stops', 'zastávky', 'počet zastávek'] if k in headers), None),
        'start_location': next((headers.get(k) for k in ['startovní místo', 'start', 'depo'] if k in headers), None),
        'distance': next((headers.get(k) for k in ['celková vzdálenost', 'vzdálenost', 'distance', 'km'] if k in headers), None),
        'work_time': next((headers.get(k) for k in ['čas práce', 'doba', 'time', 'work_time'] if k in headers), None),
        'route_type': next((headers.get(k) for k in ['dr/lh', 'typ', 'type', 'route_type'] if k in headers), None),
    }
    
    linehaul_type_detected = None
    
    # Parse routes
    for row_idx in range(header_row + 1, routes_sheet.max_row + 1):
        row = [routes_sheet.cell(row=row_idx, column=c).value for c in range(1, routes_sheet.max_column + 1)]
        
        if not row[0]:
            continue
        
        vehicle_id = str(row[col_map['vehicle_id']]) if col_map['vehicle_id'] is not None else str(row[0])
        
        route_code_match = re.search(r'[A-Z]$', vehicle_id.strip())
        route_code = route_code_match.group(0) if route_code_match else vehicle_id[-1] if vehicle_id else None
        
        route_type_raw = row[col_map['route_type']] if col_map['route_type'] is not None else 'LH'
        route_type = str(route_type_raw) if route_type_raw else 'LH'
        
        # Detect linehaul type for the batch (once)
        if linehaul_type_detected is None and route_type:
            linehaul_count = route_type.upper().count('LH')
            if linehaul_count > 0:
                result['linehauls_per_batch'] = linehaul_count
                linehaul_type_detected = route_type
        
        is_direct = 'DR' in route_type.upper() and 'LH' not in route_type.upper()
        
        if is_direct:
            result['routes_dr'] += 1
            route_type_normalized = 'DR'
        else:
            result['routes_lh'] += 1
            route_type_normalized = 'LH'
        
        # Distance
        distance_km = Decimal('0')
        if col_map['distance'] is not None and row[col_map['distance']]:
            try:
                distance_km = Decimal(str(row[col_map['distance']]))
            except:
                pass
        
        # Stops
        stops_count = 0
        if col_map['stops'] is not None and row[col_map['stops']]:
            try:
                stops_count = int(row[col_map['stops']])
            except:
                pass
        
        # Work time
        work_time_minutes = 0
        if col_map['work_time'] is not None and row[col_map['work_time']]:
            work_time = row[col_map['work_time']]
            if isinstance(work_time, time):
                work_time_minutes = work_time.hour * 60 + work_time.minute
            elif isinstance(work_time, str):
                match = re.match(r'(\d+):(\d+)', work_time)
                if match:
                    work_time_minutes = int(match.group(1)) * 60 + int(match.group(2))
        
        route_data = {
            'vehicle_id': vehicle_id,
            'route_code': route_code,
            'route_type': route_type_normalized,
            'route_type_raw': route_type,
            'start_location': row[col_map['start_location']] if col_map['start_location'] is not None else None,
            'distance_km': distance_km,
            'work_time_minutes': work_time_minutes,
            'stops_count': stops_count,
        }
        
        result['routes'].append(route_data)
        result['total_routes'] += 1
        result['total_distance_km'] += distance_km
        result['total_stops'] += stops_count
    
    return result


def aggregate_plans_for_period(plans: List[Plan]) -> dict:
    """Aggregate multiple plans into totals for comparison with proof"""
    aggregated = {
        'plan_ids': [p.id for p in plans],
        'plans_count': len(plans),
        'total_working_days': sum(p.working_days for p in plans),
        'total_routes': sum(p.total_routes for p in plans),
        'total_routes_dr': sum(p.routes_dr_per_day * p.working_days for p in plans),
        'total_routes_lh': sum(p.routes_lh_per_day * p.working_days for p in plans),
        'total_distance_km': sum(float(p.total_distance_km or 0) for p in plans),
        'routes_per_day_avg': 0,
        'linehauls_per_batch': plans[0].linehauls_per_batch if plans else 2,
    }
    
    if aggregated['total_working_days'] > 0:
        aggregated['routes_per_day_avg'] = aggregated['total_routes'] / aggregated['total_working_days']
    
    return aggregated


def compare_aggregated_plans_with_proof(aggregated: dict, proof: Proof) -> dict:
    """Compare aggregated plans with actual proof data"""
    comparison = {
        'plans_count': aggregated['plans_count'],
        'plan_ids': aggregated['plan_ids'],
        'total_working_days': aggregated['total_working_days'],
        
        'routes_planned': aggregated['total_routes'],
        'routes_actual': 0,
        'routes_difference': 0,
        
        'linehauls_per_batch': aggregated['linehauls_per_batch'],
        
        'distance_planned_km': aggregated['total_distance_km'],
        'distance_actual_km': 0,
        
        'cost_actual': float(proof.grand_total or 0),
        
        'extra_routes': 0,
        'missing_routes': 0,
        'combined_routes': 0,
        
        'differences': [],
        'route_comparison': [],
        
        'plan_breakdown': {
            'total_routes': aggregated['total_routes'],
            'routes_lh': aggregated['total_routes_lh'],
            'routes_dr': aggregated['total_routes_dr'],
            'working_days': aggregated['total_working_days'],
            'routes_per_day_avg': aggregated['routes_per_day_avg'],
            'linehauls_per_batch': aggregated['linehauls_per_batch'],
        },
        'proof_breakdown': {
            'lh_dpo': 0,
            'lh_sd': 0,
            'lh_sd_spojene': 0,
            'dr': 0,
            'total_routes': 0,
        }
    }
    
    # Parse proof route details
    for route_detail in proof.route_details:
        route_type = route_detail.route_type.upper()
        count = route_detail.count
        
        if 'DPO' in route_type:
            comparison['proof_breakdown']['lh_dpo'] += count
        elif 'SPOJENE' in route_type or 'SPOJENÉ' in route_type:
            comparison['proof_breakdown']['lh_sd_spojene'] += count
            comparison['combined_routes'] += count
        elif 'SD' in route_type:
            comparison['proof_breakdown']['lh_sd'] += count
        elif 'DR' in route_type:
            comparison['proof_breakdown']['dr'] += count
    
    comparison['routes_actual'] = (
        comparison['proof_breakdown']['lh_dpo'] +
        comparison['proof_breakdown']['lh_sd'] +
        comparison['proof_breakdown']['lh_sd_spojene'] +
        comparison['proof_breakdown']['dr']
    )
    comparison['proof_breakdown']['total_routes'] = comparison['routes_actual']
    
    comparison['routes_difference'] = comparison['routes_actual'] - comparison['routes_planned']
    
    # Route type comparison
    planned_lh = aggregated['total_routes_lh']
    actual_lh = (
        comparison['proof_breakdown']['lh_dpo'] + 
        comparison['proof_breakdown']['lh_sd'] + 
        comparison['proof_breakdown']['lh_sd_spojene']
    )
    lh_diff = actual_lh - planned_lh
    
    comparison['route_comparison'].append({
        'type': 'Last Mile trasy (LH)',
        'planned': planned_lh,
        'actual': actual_lh,
        'difference': lh_diff,
        'status': 'ok' if lh_diff == 0 else 'extra' if lh_diff > 0 else 'missing',
        'note': f'DPO: {comparison["proof_breakdown"]["lh_dpo"]}, SD: {comparison["proof_breakdown"]["lh_sd"]}, Spojené: {comparison["proof_breakdown"]["lh_sd_spojene"]}'
    })
    
    planned_dr = aggregated['total_routes_dr']
    actual_dr = comparison['proof_breakdown']['dr']
    dr_diff = actual_dr - planned_dr
    
    comparison['route_comparison'].append({
        'type': 'Direct trasy (DR)',
        'planned': planned_dr,
        'actual': actual_dr,
        'difference': dr_diff,
        'status': 'ok' if dr_diff == 0 else 'extra' if dr_diff > 0 else 'missing',
        'note': ''
    })
    
    comparison['route_comparison'].append({
        'type': 'Linehaul kamiony',
        'planned': aggregated['linehauls_per_batch'],
        'actual': aggregated['linehauls_per_batch'],
        'difference': 0,
        'status': 'info',
        'note': f'LH-LH = {aggregated["linehauls_per_batch"]} kamiony/den'
    })
    
    # Generate differences
    if lh_diff > 0:
        comparison['extra_routes'] += lh_diff
        comparison['differences'].append({
            'type': 'extra',
            'category': 'LH trasy',
            'count': lh_diff,
            'message': f'Více Last Mile tras než plánováno: +{lh_diff}'
        })
    elif lh_diff < 0:
        comparison['missing_routes'] += abs(lh_diff)
        comparison['differences'].append({
            'type': 'missing',
            'category': 'LH trasy',
            'count': abs(lh_diff),
            'message': f'Méně Last Mile tras než plánováno: {lh_diff}'
        })
    
    if dr_diff > 0:
        comparison['extra_routes'] += dr_diff
        comparison['differences'].append({
            'type': 'extra',
            'category': 'DR trasy',
            'count': dr_diff,
            'message': f'Více Direct tras než plánováno: +{dr_diff}'
        })
    elif dr_diff < 0:
        comparison['missing_routes'] += abs(dr_diff)
        comparison['differences'].append({
            'type': 'missing',
            'category': 'DR trasy',
            'count': abs(dr_diff),
            'message': f'Méně Direct tras než plánováno: {dr_diff}'
        })
    
    if comparison['combined_routes'] > 0:
        comparison['differences'].append({
            'type': 'info',
            'category': 'Efektivita',
            'count': comparison['combined_routes'],
            'message': f'Spojeno {comparison["combined_routes"]} tras (úspora {comparison["combined_routes"]} vozidel)'
        })
    
    return comparison


# =============================================================================
# API ENDPOINTS
# =============================================================================

@router.get("")
async def get_plans(
    carrier_id: Optional[int] = Query(None),
    period: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Get all plans with filters"""
    query = select(Plan).options(selectinload(Plan.carrier))
    
    filters = []
    if carrier_id:
        filters.append(Plan.carrier_id == carrier_id)
    if period:
        filters.append(Plan.period == period)
    
    if filters:
        query = query.where(and_(*filters))
    
    query = query.order_by(Plan.valid_from.desc())
    
    result = await db.execute(query)
    plans = result.scalars().all()
    
    return [{
        'id': p.id,
        'carrierId': p.carrier_id,
        'carrierName': p.carrier.name if p.carrier else None,
        'name': p.name,
        'validFrom': p.valid_from.isoformat(),
        'validTo': p.valid_to.isoformat(),
        'period': p.period,
        'workingDays': p.working_days,
        'routesPerDay': p.routes_per_day,
        'totalRoutes': p.total_routes,
        'totalDistanceKm': float(p.total_distance_km) if p.total_distance_km else 0,
        'linehaulsPerBatch': p.linehauls_per_batch,
        'routesDrPerDay': p.routes_dr_per_day,
        'routesLhPerDay': p.routes_lh_per_day,
        'status': p.status,
        'createdAt': p.created_at.isoformat() if p.created_at else None,
    } for p in plans]


@router.post("/upload", status_code=201)
async def upload_plan(
    file: UploadFile = File(...),
    carrier_id: int = Form(...),
    valid_from: str = Form(...),  # Only require start date
    db: AsyncSession = Depends(get_db)
):
    """Upload and parse planning XLSX file
    
    valid_to is automatically calculated:
    - If there's a next plan, valid_to = next_plan.valid_from - 1 day
    - If no next plan, valid_to = end of month
    
    When uploading a new plan, previous plan's valid_to is also updated.
    """
    # Validate carrier
    carrier_result = await db.execute(
        select(Carrier).where(Carrier.id == carrier_id)
    )
    carrier = carrier_result.scalar_one_or_none()
    if not carrier:
        raise HTTPException(status_code=400, detail="Carrier not found")
    
    # Parse date
    try:
        parsed_from = date.fromisoformat(valid_from)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    period = parsed_from.strftime("%m/%Y")
    
    # Calculate valid_to as end of month (will be adjusted if another plan is uploaded)
    if parsed_from.month == 12:
        end_of_month = date(parsed_from.year + 1, 1, 1) - timedelta(days=1)
    else:
        end_of_month = date(parsed_from.year, parsed_from.month + 1, 1) - timedelta(days=1)
    
    parsed_to = end_of_month
    
    # Check for existing plans in this period and update their valid_to if needed
    existing_plans_result = await db.execute(
        select(Plan)
        .where(and_(
            Plan.carrier_id == carrier_id,
            Plan.period == period,
            Plan.valid_from < parsed_from  # Plans that start before this one
        ))
        .order_by(Plan.valid_from.desc())
    )
    previous_plan = existing_plans_result.scalars().first()
    
    if previous_plan:
        # Update previous plan's valid_to to day before this plan starts
        previous_plan.valid_to = parsed_from - timedelta(days=1)
        previous_plan.working_days = count_working_days(previous_plan.valid_from, previous_plan.valid_to)
        previous_plan.total_routes = previous_plan.routes_per_day * previous_plan.working_days
        if previous_plan.distance_per_day_km:
            previous_plan.total_distance_km = previous_plan.distance_per_day_km * previous_plan.working_days
    
    # Check for plans that start after this one (to set our valid_to)
    next_plans_result = await db.execute(
        select(Plan)
        .where(and_(
            Plan.carrier_id == carrier_id,
            Plan.period == period,
            Plan.valid_from > parsed_from
        ))
        .order_by(Plan.valid_from.asc())
    )
    next_plan = next_plans_result.scalars().first()
    
    if next_plan:
        # Our valid_to is day before next plan starts
        parsed_to = next_plan.valid_from - timedelta(days=1)
    
    working_days = count_working_days(parsed_from, parsed_to)
    
    # Parse file
    content = await file.read()
    
    try:
        plan_data = parse_plan_from_xlsx(content, parsed_from)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {str(e)}")
    
    # Create plan
    plan = Plan(
        carrier_id=carrier_id,
        name=file.filename or f"Plan od {valid_from}",
        valid_from=parsed_from,
        valid_to=parsed_to,
        period=period,
        file_name=file.filename,
        routes_per_day=plan_data['total_routes'],
        linehauls_per_batch=plan_data['linehauls_per_batch'],
        distance_per_day_km=plan_data['total_distance_km'],
        stops_per_day=plan_data['total_stops'],
        routes_dr_per_day=plan_data['routes_dr'],
        routes_lh_per_day=plan_data['routes_lh'],
        working_days=working_days,
        total_routes=plan_data['total_routes'] * working_days,
        total_distance_km=plan_data['total_distance_km'] * working_days if plan_data['total_distance_km'] else None,
        status='active'
    )
    db.add(plan)
    await db.flush()
    
    for route in plan_data['routes']:
        db.add(PlanRoute(
            plan_id=plan.id,
            vehicle_id=route['vehicle_id'],
            route_code=route['route_code'],
            route_type=route['route_type'],
            start_location=route['start_location'],
            distance_km=route['distance_km'],
            work_time_minutes=route['work_time_minutes'],
            stops_count=route['stops_count'],
        ))
    
    await db.commit()
    
    # Prepare info about updated previous plan
    updated_previous = None
    if previous_plan:
        updated_previous = {
            'id': previous_plan.id,
            'name': previous_plan.name,
            'validTo': previous_plan.valid_to.isoformat(),
            'workingDays': previous_plan.working_days,
        }
    
    return {
        'success': True,
        'message': 'Plan uploaded successfully',
        'data': {
            'id': plan.id,
            'name': plan.name,
            'validFrom': plan.valid_from.isoformat(),
            'validTo': plan.valid_to.isoformat(),
            'validToNote': 'Konec měsíce' if parsed_to == end_of_month else 'Den před dalším plánem',
            'period': plan.period,
            'workingDays': plan.working_days,
            'routesPerDay': plan.routes_per_day,
            'totalRoutes': plan.total_routes,
            'totalDistanceKm': float(plan.total_distance_km) if plan.total_distance_km else 0,
            'linehaulsPerBatch': plan.linehauls_per_batch,
            'routesDrPerDay': plan.routes_dr_per_day,
            'routesLhPerDay': plan.routes_lh_per_day,
            'routes': [{
                'vehicleId': r['vehicle_id'],
                'routeCode': r['route_code'],
                'routeType': r['route_type'],
                'routeTypeRaw': r['route_type_raw'],
                'distanceKm': float(r['distance_km']),
                'stopsCount': r['stops_count'],
            } for r in plan_data['routes']]
        },
        'updatedPreviousPlan': updated_previous
    }


@router.get("/by-period/{period}")
async def get_plans_by_period(
    period: str,
    carrier_id: int = Query(...),
    db: AsyncSession = Depends(get_db)
):
    """Get all plans for a specific period (MM/YYYY) and carrier"""
    result = await db.execute(
        select(Plan)
        .options(selectinload(Plan.carrier))
        .where(and_(
            Plan.carrier_id == carrier_id,
            Plan.period == period
        ))
        .order_by(Plan.valid_from)
    )
    plans = result.scalars().all()
    
    return [{
        'id': p.id,
        'name': p.name,
        'validFrom': p.valid_from.isoformat(),
        'validTo': p.valid_to.isoformat(),
        'workingDays': p.working_days,
        'routesPerDay': p.routes_per_day,
        'totalRoutes': p.total_routes,
        'linehaulsPerBatch': p.linehauls_per_batch,
    } for p in plans]


@router.get("/{plan_id}")
async def get_plan(plan_id: int, db: AsyncSession = Depends(get_db)):
    """Get plan details with routes"""
    result = await db.execute(
        select(Plan)
        .options(selectinload(Plan.carrier), selectinload(Plan.routes))
        .where(Plan.id == plan_id)
    )
    plan = result.scalar_one_or_none()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return {
        'id': plan.id,
        'carrierId': plan.carrier_id,
        'carrierName': plan.carrier.name if plan.carrier else None,
        'name': plan.name,
        'validFrom': plan.valid_from.isoformat(),
        'validTo': plan.valid_to.isoformat(),
        'period': plan.period,
        'workingDays': plan.working_days,
        'routesPerDay': plan.routes_per_day,
        'totalRoutes': plan.total_routes,
        'totalDistanceKm': float(plan.total_distance_km) if plan.total_distance_km else 0,
        'linehaulsPerBatch': plan.linehauls_per_batch,
        'routesDrPerDay': plan.routes_dr_per_day,
        'routesLhPerDay': plan.routes_lh_per_day,
        'status': plan.status,
        'routes': [{
            'id': r.id,
            'vehicleId': r.vehicle_id,
            'routeCode': r.route_code,
            'routeType': r.route_type,
            'distanceKm': float(r.distance_km) if r.distance_km else 0,
            'stopsCount': r.stops_count,
        } for r in plan.routes]
    }


@router.post("/compare-with-proof/{proof_id}")
async def compare_plans_with_proof(
    proof_id: int,
    plan_ids: List[int] = Query(...),
    db: AsyncSession = Depends(get_db)
):
    """Compare multiple plans with a proof"""
    # Get proof
    proof_result = await db.execute(
        select(Proof)
        .options(
            selectinload(Proof.carrier),
            selectinload(Proof.route_details)
        )
        .where(Proof.id == proof_id)
    )
    proof = proof_result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    # Get plans
    plans_result = await db.execute(
        select(Plan)
        .options(selectinload(Plan.carrier))
        .where(Plan.id.in_(plan_ids))
        .order_by(Plan.valid_from)
    )
    plans = plans_result.scalars().all()
    
    if not plans:
        raise HTTPException(status_code=400, detail="No plans found")
    
    # Verify all plans are for same carrier as proof
    for plan in plans:
        if plan.carrier_id != proof.carrier_id:
            raise HTTPException(
                status_code=400, 
                detail=f"Plan {plan.id} is for different carrier"
            )
    
    # Aggregate plans and compare
    aggregated = aggregate_plans_for_period(plans)
    comparison = compare_aggregated_plans_with_proof(aggregated, proof)
    
    # Save comparison
    import json
    plan_comparison = PlanComparison(
        proof_id=proof_id,
        plan_ids=','.join(str(pid) for pid in plan_ids),
        plans_count=len(plans),
        total_working_days=aggregated['total_working_days'],
        routes_planned=aggregated['total_routes'],
        routes_actual=comparison['routes_actual'],
        routes_difference=comparison['routes_difference'],
        distance_planned_km=Decimal(str(aggregated['total_distance_km'])),
        cost_actual=Decimal(str(comparison['cost_actual'])),
        extra_routes=comparison['extra_routes'],
        missing_routes=comparison['missing_routes'],
        combined_routes=comparison['combined_routes'],
        differences_json=json.dumps(comparison['differences']),
        status='completed'
    )
    db.add(plan_comparison)
    await db.commit()
    
    return {
        'success': True,
        'comparison': comparison,
        'plans': [{
            'id': p.id,
            'name': p.name,
            'validFrom': p.valid_from.isoformat(),
            'validTo': p.valid_to.isoformat(),
            'workingDays': p.working_days,
            'totalRoutes': p.total_routes,
        } for p in plans],
        'proof': {
            'id': proof.id,
            'period': proof.period,
            'grandTotal': float(proof.grand_total) if proof.grand_total else 0,
        }
    }


@router.delete("/{plan_id}", status_code=204)
async def delete_plan(plan_id: int, db: AsyncSession = Depends(get_db)):
    """Delete plan"""
    result = await db.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    await db.delete(plan)
    await db.commit()
