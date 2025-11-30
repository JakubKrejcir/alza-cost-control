"""
Route Plans API Router - with XLSX upload and parsing
UPDATED: Support for plan_type (BOTH/DPO/SD) based on filename suffix
UPDATED: Aggregated comparison - proof vs all plans valid in the month
"""
from typing import List, Optional
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
import re
import calendar
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import openpyxl

from app.database import get_db
from app.models import RoutePlan, RoutePlanRoute, RoutePlanDetail, Carrier, Proof

router = APIRouter()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_time_to_hour(time_val) -> Optional[int]:
    """Extract hour from time value"""
    if time_val is None:
        return None
    
    if isinstance(time_val, str):
        match = re.match(r'(\d{1,2}):', time_val)
        if match:
            return int(match.group(1))
    elif hasattr(time_val, 'hour'):
        return time_val.hour
    
    return None


def determine_route_type(start_time) -> str:
    """Determine if route is DPO (morning) or SD (afternoon) based on start time"""
    hour = parse_time_to_hour(start_time)
    if hour is None:
        return "DPO"  # Default
    
    # Before 12:00 = DPO (morning), after = SD (afternoon)
    return "DPO" if hour < 12 else "SD"


def extract_route_letter(route_name: str) -> Optional[str]:
    """Extract letter from route name like 'Moravskoslezsko A' -> 'A'"""
    if not route_name:
        return None
    parts = route_name.strip().split()
    if parts:
        return parts[-1]
    return None


def parse_linehaul_count(delivery_type: str) -> int:
    """Parse LH count from delivery type like 'LH-LH' -> 2"""
    if not delivery_type:
        return 0
    
    # Count occurrences of 'LH'
    return delivery_type.upper().count('LH')


def detect_plan_type_from_filename(filename: str) -> str:
    """
    Detect plan type from filename suffix.
    
    - filename_DPO.xlsx -> DPO (ranní rozvozy)
    - filename_SD.xlsx -> SD (odpolední rozvozy)
    - filename.xlsx -> BOTH (platí pro obojí)
    """
    if not filename:
        return "BOTH"
    
    # Remove extension and check suffix
    name_without_ext = filename.rsplit('.', 1)[0]
    name_upper = name_without_ext.upper()
    
    if name_upper.endswith('_DPO'):
        return "DPO"
    elif name_upper.endswith('_SD'):
        return "SD"
    else:
        return "BOTH"


def parse_date_from_filename(filename: str) -> Optional[datetime]:
    """
    Extract date from filename.
    
    Supported formats:
    - Drivecool 25-08-22.xlsx -> 2025-08-22 (YY-MM-DD)
    - plan_01.12.2025.xlsx -> 2025-12-01 (DD.MM.YYYY)
    - plan_01-12-2025.xlsx -> 2025-12-01 (DD-MM-YYYY)
    - plan_01.12.25.xlsx -> 2025-12-01 (DD.MM.YY)
    """
    
    # First try to match 4-digit year patterns (more specific)
    patterns_4digit = [
        (r'(\d{2})\.(\d{2})\.(\d{4})', 'DD.MM.YYYY'),
        (r'(\d{2})-(\d{2})-(\d{4})', 'DD-MM-YYYY'),
        (r'(\d{2})_(\d{2})_(\d{4})', 'DD_MM_YYYY'),
        (r'(\d{4})\.(\d{2})\.(\d{2})', 'YYYY.MM.DD'),
        (r'(\d{4})-(\d{2})-(\d{2})', 'YYYY-MM-DD'),
        (r'(\d{4})_(\d{2})_(\d{2})', 'YYYY_MM_DD'),
    ]
    
    for pattern, fmt in patterns_4digit:
        match = re.search(pattern, filename)
        if match:
            g0, g1, g2 = int(match.group(1)), int(match.group(2)), int(match.group(3))
            
            if fmt.startswith('YYYY'):
                year, month, day = g0, g1, g2
            else:
                day, month, year = g0, g1, g2
            
            try:
                return datetime(year, month, day)
            except ValueError:
                continue
    
    # Now try 2-digit year patterns
    patterns_2digit = [
        r'(\d{2})\.(\d{2})\.(\d{2})',
        r'(\d{2})-(\d{2})-(\d{2})',
        r'(\d{2})_(\d{2})_(\d{2})',
    ]
    
    for pattern in patterns_2digit:
        match = re.search(pattern, filename)
        if match:
            g0, g1, g2 = int(match.group(1)), int(match.group(2)), int(match.group(3))
            
            # Heuristics to determine format:
            # If first number is 20-99, it's likely YY (year 2020-2099)
            # If first number is > 12 and <= 31, it's likely DD
            
            # Check if it looks like YY-MM-DD (year first)
            # Years 20-30 are common for current decade
            if g0 >= 20 and g0 <= 35 and g1 >= 1 and g1 <= 12 and g2 >= 1 and g2 <= 31:
                # YY-MM-DD format (e.g., 25-08-22 = 2025-08-22)
                year = 2000 + g0
                month = g1
                day = g2
            elif g0 > 12 and g0 <= 31:
                # DD-MM-YY (day > 12, so first is day)
                day, month = g0, g1
                year = 2000 + g2
            elif g2 > 12 and g2 <= 31:
                # YY-MM-DD (last > 12, so last is day)
                year = 2000 + g0
                month = g1
                day = g2
            else:
                # Default: assume DD-MM-YY (Czech format)
                day, month = g0, g1
                year = 2000 + g2
            
            try:
                return datetime(year, month, day)
            except ValueError:
                # If failed, try YY-MM-DD interpretation
                try:
                    return datetime(2000 + g0, g1, g2)
                except ValueError:
                    continue
    
    return None


def get_working_days_in_range(start_date: datetime, end_date: datetime) -> int:
    """Count working days (Mon-Fri) between two dates inclusive"""
    if not start_date or not end_date:
        return 0
    
    count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # Monday = 0, Friday = 4
            count += 1
        current += timedelta(days=1)
    return count


def parse_route_plan_xlsx(file_content: bytes, filename: str) -> dict:
    """Parse route plan data from XLSX"""
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    
    result = {
        'valid_from': parse_date_from_filename(filename),
        'plan_type': detect_plan_type_from_filename(filename),
        'routes': [],
        'details': [],  # For future use
        'summary': {
            'total_routes': 0,
            'dpo_routes_count': 0,
            'sd_routes_count': 0,
            'dpo_linehaul_count': 0,
            'sd_linehaul_count': 0,
            'total_distance_km': Decimal('0'),
            'total_stops': 0,
        }
    }
    
    # Parse Routes sheet
    if 'Routes' in wb.sheetnames:
        sheet = wb['Routes']
        
        # Find header row
        headers = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers[val.strip()] = col
        
        # Map expected columns
        col_map = {
            'route_name': headers.get('Identifikátor vozidla'),
            'carrier': headers.get('Dopravce'),
            'stops': headers.get('Náklad 1'),
            'start_location': headers.get('Startovní místo'),
            'max_capacity': headers.get('Max Náklad 2'),
            'start_time': headers.get('Začátek'),
            'end_time': headers.get('Konec'),
            'distance': headers.get('Celková vzdálenost'),
            'work_time': headers.get('Čas práce'),
            'delivery_type': headers.get('DR/LH'),
        }
        
        dpo_lh_set = set()
        sd_lh_set = set()
        
        for row in range(2, sheet.max_row + 1):
            route_name = sheet.cell(row=row, column=col_map['route_name']).value if col_map['route_name'] else None
            if not route_name:
                continue
            
            def time_to_str(t):
                if t is None:
                    return None
                if isinstance(t, str):
                    return t
                if hasattr(t, 'strftime'):
                    return t.strftime('%H:%M')
                return str(t)
            
            start_time = sheet.cell(row=row, column=col_map['start_time']).value if col_map['start_time'] else None
            start_time_str = time_to_str(start_time)
            route_type = determine_route_type(start_time)
            
            delivery_type = sheet.cell(row=row, column=col_map['delivery_type']).value if col_map['delivery_type'] else None
            delivery_type_str = str(delivery_type).strip().upper() if delivery_type else None
            
            # Count linehauls per batch (LH-LH = 2 kamiony)
            if delivery_type_str and 'LH' in delivery_type_str:
                lh_count = parse_linehaul_count(delivery_type_str)
                if route_type == 'DPO':
                    dpo_lh_set.add(delivery_type_str)
                else:
                    sd_lh_set.add(delivery_type_str)
            
            stops_count = int(sheet.cell(row=row, column=col_map['stops']).value or 0) if col_map['stops'] else 0
            distance = Decimal(str(sheet.cell(row=row, column=col_map['distance']).value or 0)) if col_map['distance'] else Decimal('0')
            
            route_data = {
                'route_name': str(route_name),
                'route_letter': extract_route_letter(str(route_name)),
                'carrier_name': sheet.cell(row=row, column=col_map['carrier']).value if col_map['carrier'] else None,
                'route_type': route_type,
                'delivery_type': delivery_type_str,
                'start_location': sheet.cell(row=row, column=col_map['start_location']).value if col_map['start_location'] else None,
                'stops_count': stops_count,
                'max_capacity': int(sheet.cell(row=row, column=col_map['max_capacity']).value or 0) if col_map['max_capacity'] else 0,
                'start_time': start_time_str,
                'end_time': time_to_str(sheet.cell(row=row, column=col_map['end_time']).value if col_map['end_time'] else None),
                'work_time': time_to_str(sheet.cell(row=row, column=col_map['work_time']).value if col_map['work_time'] else None),
                'distance_km': distance,
            }
            
            result['routes'].append(route_data)
            result['summary']['total_routes'] += 1
            result['summary']['total_stops'] += stops_count
            result['summary']['total_distance_km'] += distance
            
            if route_type == 'DPO':
                result['summary']['dpo_routes_count'] += 1
            else:
                result['summary']['sd_routes_count'] += 1
        
        # Linehaul count = 2 per batch (LH-LH = 2 kamiony pro celý batch)
        result['summary']['dpo_linehaul_count'] = 2 if dpo_lh_set else 0
        result['summary']['sd_linehaul_count'] = 2 if sd_lh_set else 0
    
    return result


async def get_route_plan_by_id(plan_id: int, db: AsyncSession) -> Optional[RoutePlan]:
    """Get route plan with routes loaded"""
    result = await db.execute(
        select(RoutePlan)
        .options(
            selectinload(RoutePlan.carrier),
            selectinload(RoutePlan.routes)
        )
        .where(RoutePlan.id == plan_id)
    )
    return result.scalar_one_or_none()


async def update_route_plan_validity(carrier_id: int, db: AsyncSession, plan_type: str = None):
    """
    Update valid_to dates for all plans of a carrier.
    
    Logic: Plans are grouped by validFrom (period start). 
    All plans in a period end when the next period starts.
    
    Example:
    - Period 1 (2025-10-10): DPO + SD plans -> valid_to = 2025-10-23
    - Period 2 (2025-10-24): BOTH plan -> valid_to = None (current)
    """
    # Get all unique validFrom dates for this carrier
    result = await db.execute(
        select(RoutePlan.valid_from)
        .where(RoutePlan.carrier_id == carrier_id)
        .distinct()
        .order_by(RoutePlan.valid_from.asc())
    )
    periods = [row[0] for row in result.fetchall()]
    
    # For each period, set valid_to based on next period
    for i, period_start in enumerate(periods):
        # Get all plans for this period
        plans_result = await db.execute(
            select(RoutePlan).where(
                and_(
                    RoutePlan.carrier_id == carrier_id,
                    RoutePlan.valid_from == period_start
                )
            )
        )
        plans = plans_result.scalars().all()
        
        # Determine valid_to
        if i + 1 < len(periods):
            # There's a next period - end day before it starts
            valid_to = periods[i + 1] - timedelta(days=1)
        else:
            # Last period - no end date
            valid_to = None
        
        # Update all plans in this period
        for plan in plans:
            plan.valid_to = valid_to


async def delete_plans_for_date(
    carrier_id: int, 
    valid_from: datetime, 
    db: AsyncSession,
    exclude_plan_type: Optional[str] = None
) -> int:
    """
    Delete all plans for a specific carrier and date.
    
    Args:
        carrier_id: Carrier ID
        valid_from: The date (period start)
        db: Database session
        exclude_plan_type: If set, don't delete plans of this type (for partial updates)
    
    Returns:
        Number of deleted plans
    """
    query = select(RoutePlan).where(
        and_(
            RoutePlan.carrier_id == carrier_id,
            RoutePlan.valid_from == valid_from
        )
    )
    
    if exclude_plan_type:
        query = query.where(RoutePlan.plan_type != exclude_plan_type)
    
    result = await db.execute(query)
    existing_plans = result.scalars().all()
    
    for plan in existing_plans:
        await db.delete(plan)
    
    await db.flush()
    return len(existing_plans)


async def check_plan_type_conflicts(
    carrier_id: int, 
    valid_from: datetime, 
    plan_type: str, 
    db: AsyncSession
) -> Optional[str]:
    """
    Handle plan type conflicts by deleting conflicting plans.
    
    Rules:
    - BOTH = complete plan, replaces everything for that date
    - DPO/SD = partial plans, can coexist with each other but replace BOTH
    
    Returns error message only for unexpected issues, None otherwise.
    """
    # Find existing plans for same carrier and date
    result = await db.execute(
        select(RoutePlan).where(
            and_(
                RoutePlan.carrier_id == carrier_id,
                RoutePlan.valid_from == valid_from
            )
        )
    )
    existing_plans = result.scalars().all()
    
    for existing in existing_plans:
        # BOTH replaces everything
        if plan_type == "BOTH":
            await db.delete(existing)
            continue
        
        # Partial plan (DPO/SD/etc.) replaces BOTH
        if existing.plan_type == "BOTH":
            await db.delete(existing)
            continue
        
        # Same type - will be replaced later in upload logic
        if plan_type == existing.plan_type:
            continue
        
        # Different partial types can coexist (DPO + SD, etc.)
    
    await db.flush()
    return None


async def get_plans_for_period(
    carrier_id: int, 
    period: str, 
    db: AsyncSession
) -> List[RoutePlan]:
    """
    Get all plans valid during a calendar month period.
    
    A plan is valid for a period if:
    - plan.valid_from <= last day of period AND
    - (plan.valid_to IS NULL OR plan.valid_to >= first day of period)
    """
    # Parse period (MM/YYYY)
    try:
        month, year = period.split('/')
        month, year = int(month), int(year)
    except ValueError:
        return []
    
    # Get first and last day of month
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, calendar.monthrange(year, month)[1])
    
    # Query plans that overlap with this period
    result = await db.execute(
        select(RoutePlan)
        .options(selectinload(RoutePlan.routes))
        .where(
            and_(
                RoutePlan.carrier_id == carrier_id,
                RoutePlan.valid_from <= last_day,
                or_(
                    RoutePlan.valid_to == None,
                    RoutePlan.valid_to >= first_day
                )
            )
        )
        .order_by(RoutePlan.valid_from.asc())
    )
    return result.scalars().all()


def aggregate_plans_for_period(
    plans: List[RoutePlan], 
    period: str
) -> dict:
    """
    Aggregate multiple plans for a period, calculating weighted values.
    
    Each plan contributes based on how many working days it covers in the period.
    """
    # Parse period
    try:
        month, year = period.split('/')
        month, year = int(month), int(year)
    except ValueError:
        return None
    
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, calendar.monthrange(year, month)[1])
    
    # Total working days in month
    total_working_days = get_working_days_in_range(first_day, last_day)
    
    if total_working_days == 0:
        return None
    
    aggregated = {
        'plans': [],
        'totalWorkingDays': total_working_days,
        'dpoRoutesCount': 0,
        'sdRoutesCount': 0,
        'dpoLinehaulCount': 0,
        'sdLinehaulCount': 0,
        'totalRoutes': 0,
        'coverage': [],
        'missingDays': 0,
    }
    
    covered_days = set()
    
    for plan in plans:
        # Calculate overlap with period
        plan_start = max(plan.valid_from, first_day)
        plan_end = min(plan.valid_to, last_day) if plan.valid_to else last_day
        
        if plan_start > plan_end:
            continue  # No overlap
        
        working_days = get_working_days_in_range(plan_start, plan_end)
        weight = working_days / total_working_days
        
        # Track covered days
        current = plan_start
        while current <= plan_end:
            if current.weekday() < 5:
                covered_days.add(current.date())
            current += timedelta(days=1)
        
        plan_info = {
            'id': plan.id,
            'planType': plan.plan_type,
            'fileName': plan.file_name,
            'validFrom': plan.valid_from.isoformat(),
            'validTo': plan.valid_to.isoformat() if plan.valid_to else None,
            'overlapStart': plan_start.isoformat(),
            'overlapEnd': plan_end.isoformat(),
            'workingDays': working_days,
            'weight': round(weight, 3),
            'dpoRoutesCount': plan.dpo_routes_count,
            'sdRoutesCount': plan.sd_routes_count,
            'dpoLinehaulCount': plan.dpo_linehaul_count,
            'sdLinehaulCount': plan.sd_linehaul_count,
        }
        aggregated['plans'].append(plan_info)
        
        # Weighted contribution
        # Routes: multiply by working days (routes run every day)
        if plan.plan_type in ('BOTH', 'DPO'):
            aggregated['dpoRoutesCount'] += plan.dpo_routes_count * working_days
        if plan.plan_type in ('BOTH', 'SD'):
            aggregated['sdRoutesCount'] += plan.sd_routes_count * working_days
        
        # Linehaul: multiply by working days (linehaul runs every day too)
        if plan.plan_type in ('BOTH', 'DPO'):
            aggregated['dpoLinehaulCount'] += plan.dpo_linehaul_count * working_days
        if plan.plan_type in ('BOTH', 'SD'):
            aggregated['sdLinehaulCount'] += plan.sd_linehaul_count * working_days
    
    # Calculate missing days
    all_working_days = set()
    current = first_day
    while current <= last_day:
        if current.weekday() < 5:
            all_working_days.add(current.date())
        current += timedelta(days=1)
    
    missing_days = all_working_days - covered_days
    aggregated['missingDays'] = len(missing_days)
    aggregated['missingDates'] = sorted([d.isoformat() for d in missing_days])
    
    aggregated['totalRoutes'] = aggregated['dpoRoutesCount'] + aggregated['sdRoutesCount']
    
    return aggregated


# =============================================================================
# API ENDPOINTS
# =============================================================================

@router.get("")
async def get_route_plans(
    carrier_id: Optional[int] = Query(None),
    period: Optional[str] = Query(None),
    plan_type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Get all route plans with optional filters"""
    query = select(RoutePlan).options(selectinload(RoutePlan.carrier))
    
    filters = []
    if carrier_id:
        filters.append(RoutePlan.carrier_id == carrier_id)
    if plan_type:
        filters.append(RoutePlan.plan_type == plan_type)
    
    if filters:
        query = query.where(and_(*filters))
    
    query = query.order_by(RoutePlan.valid_from.desc())
    
    result = await db.execute(query)
    plans = result.scalars().all()
    
    return [
        {
            'id': p.id,
            'carrierId': p.carrier_id,
            'carrierName': p.carrier.name if p.carrier else None,
            'validFrom': p.valid_from.isoformat() if p.valid_from else None,
            'validTo': p.valid_to.isoformat() if p.valid_to else None,
            'fileName': p.file_name,
            'planType': p.plan_type,
            'totalRoutes': p.total_routes,
            'dpoRoutesCount': p.dpo_routes_count,
            'sdRoutesCount': p.sd_routes_count,
            'dpoLinehaulCount': p.dpo_linehaul_count,
            'sdLinehaulCount': p.sd_linehaul_count,
            'totalDistanceKm': float(p.total_distance_km) if p.total_distance_km else 0,
            'totalStops': p.total_stops,
        }
        for p in plans
    ]


@router.post("/upload")
async def upload_route_plan(
    file: UploadFile = File(...),
    carrier_id: int = Form(...),
    valid_from: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    """Upload and parse route plan XLSX"""
    # Validate carrier
    carrier_result = await db.execute(
        select(Carrier).where(Carrier.id == carrier_id)
    )
    if not carrier_result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Carrier not found")
    
    content = await file.read()
    
    try:
        plan_data = parse_route_plan_xlsx(content, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {str(e)}")
    
    # Determine valid_from
    if valid_from:
        try:
            parsed_date = datetime.fromisoformat(valid_from)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    elif plan_data['valid_from']:
        parsed_date = plan_data['valid_from']
    else:
        raise HTTPException(status_code=400, detail="Could not detect date from filename. Please provide valid_from parameter.")
    
    plan_type = plan_data['plan_type']
    
    # Check for conflicts
    conflict_error = await check_plan_type_conflicts(carrier_id, parsed_date, plan_type, db)
    if conflict_error:
        raise HTTPException(status_code=409, detail=conflict_error)
    
    # Check for existing plan with same type and date - replace it
    existing_result = await db.execute(
        select(RoutePlan).where(
            and_(
                RoutePlan.carrier_id == carrier_id,
                RoutePlan.valid_from == parsed_date,
                RoutePlan.plan_type == plan_type
            )
        )
    )
    existing = existing_result.scalar_one_or_none()
    
    if existing:
        # Delete existing plan (cascade deletes routes)
        await db.delete(existing)
        await db.flush()
    
    # Create new plan
    route_plan = RoutePlan(
        carrier_id=carrier_id,
        valid_from=parsed_date,
        file_name=file.filename,
        plan_type=plan_type,
        total_routes=plan_data['summary']['total_routes'],
        dpo_routes_count=plan_data['summary']['dpo_routes_count'],
        sd_routes_count=plan_data['summary']['sd_routes_count'],
        dpo_linehaul_count=plan_data['summary']['dpo_linehaul_count'],
        sd_linehaul_count=plan_data['summary']['sd_linehaul_count'],
        total_distance_km=plan_data['summary']['total_distance_km'],
        total_stops=plan_data['summary']['total_stops'],
    )
    db.add(route_plan)
    await db.flush()
    
    # Create routes
    route_map = {}
    for route_data in plan_data['routes']:
        route = RoutePlanRoute(
            route_plan_id=route_plan.id,
            route_name=route_data['route_name'],
            route_letter=route_data['route_letter'],
            carrier_name=route_data['carrier_name'],
            route_type=route_data['route_type'],
            delivery_type=route_data['delivery_type'],
            start_location=route_data['start_location'],
            stops_count=route_data['stops_count'],
            max_capacity=route_data['max_capacity'],
            start_time=route_data['start_time'],
            end_time=route_data['end_time'],
            work_time=route_data['work_time'],
            distance_km=route_data['distance_km'],
        )
        db.add(route)
        await db.flush()
        route_map[route_data['route_name']] = route.id
    
    # Update validity of all plans for this carrier and plan type
    await update_route_plan_validity(carrier_id, db, plan_type)
    
    await db.commit()
    
    return {
        'success': True,
        'message': f'Plánovací soubor úspěšně nahrán jako {plan_type}',
        'data': {
            'id': route_plan.id,
            'planType': route_plan.plan_type,
            'validFrom': route_plan.valid_from.isoformat(),
            'validTo': route_plan.valid_to.isoformat() if route_plan.valid_to else None,
            'totalRoutes': route_plan.total_routes,
            'dpoRoutesCount': route_plan.dpo_routes_count,
            'sdRoutesCount': route_plan.sd_routes_count,
            'dpoLinehaulCount': route_plan.dpo_linehaul_count,
            'sdLinehaulCount': route_plan.sd_linehaul_count,
            'totalDistanceKm': float(route_plan.total_distance_km) if route_plan.total_distance_km else 0,
            'totalStops': route_plan.total_stops,
        }
    }


@router.post("/upload-batch")
async def upload_route_plans_batch(
    files: List[UploadFile] = File(...),
    carrier_id: int = Form(...),
    db: AsyncSession = Depends(get_db)
):
    """Upload multiple route plan XLSX files at once"""
    from typing import List as TypingList
    
    # Validate carrier
    carrier_result = await db.execute(
        select(Carrier).where(Carrier.id == carrier_id)
    )
    if not carrier_result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Carrier not found")
    
    results = []
    errors = []
    plan_types_updated = set()
    
    for file in files:
        try:
            content = await file.read()
            plan_data = parse_route_plan_xlsx(content, file.filename)
            
            if not plan_data['valid_from']:
                errors.append({
                    'fileName': file.filename,
                    'error': 'Nepodařilo se detekovat datum z názvu souboru'
                })
                continue
            
            parsed_date = plan_data['valid_from']
            plan_type = plan_data['plan_type']
            
            # Check for conflicts
            conflict_error = await check_plan_type_conflicts(carrier_id, parsed_date, plan_type, db)
            if conflict_error:
                errors.append({
                    'fileName': file.filename,
                    'error': conflict_error
                })
                continue
            
            # Check for existing plan with same type and date - replace it
            existing_result = await db.execute(
                select(RoutePlan).where(
                    and_(
                        RoutePlan.carrier_id == carrier_id,
                        RoutePlan.valid_from == parsed_date,
                        RoutePlan.plan_type == plan_type
                    )
                )
            )
            existing = existing_result.scalar_one_or_none()
            
            if existing:
                await db.delete(existing)
                await db.flush()
            
            # Create new plan
            route_plan = RoutePlan(
                carrier_id=carrier_id,
                valid_from=parsed_date,
                file_name=file.filename,
                plan_type=plan_type,
                total_routes=plan_data['summary']['total_routes'],
                dpo_routes_count=plan_data['summary']['dpo_routes_count'],
                sd_routes_count=plan_data['summary']['sd_routes_count'],
                dpo_linehaul_count=plan_data['summary']['dpo_linehaul_count'],
                sd_linehaul_count=plan_data['summary']['sd_linehaul_count'],
                total_distance_km=plan_data['summary']['total_distance_km'],
                total_stops=plan_data['summary']['total_stops'],
            )
            db.add(route_plan)
            await db.flush()
            
            # Create routes
            for route_data in plan_data['routes']:
                route = RoutePlanRoute(
                    route_plan_id=route_plan.id,
                    route_name=route_data['route_name'],
                    route_letter=route_data['route_letter'],
                    carrier_name=route_data['carrier_name'],
                    route_type=route_data['route_type'],
                    delivery_type=route_data['delivery_type'],
                    start_location=route_data['start_location'],
                    stops_count=route_data['stops_count'],
                    max_capacity=route_data['max_capacity'],
                    start_time=route_data['start_time'],
                    end_time=route_data['end_time'],
                    work_time=route_data['work_time'],
                    distance_km=route_data['distance_km'],
                )
                db.add(route)
            
            plan_types_updated.add(plan_type)
            
            results.append({
                'fileName': file.filename,
                'id': route_plan.id,
                'planType': plan_type,
                'validFrom': parsed_date.isoformat(),
                'totalRoutes': route_plan.total_routes,
            })
            
        except Exception as e:
            # Rollback the failed transaction to continue with next file
            await db.rollback()
            errors.append({
                'fileName': file.filename,
                'error': str(e)
            })
    
    # Update validity for all affected plan types
    if results:  # Only if we have successful uploads
        for plan_type in plan_types_updated:
            await update_route_plan_validity(carrier_id, db, plan_type)
        
        await db.commit()
    
    return {
        'success': len(errors) == 0,
        'message': f'Nahráno {len(results)} souborů' + (f', {len(errors)} selhalo' if errors else ''),
        'uploaded': results,
        'errors': errors
    }


@router.get("/compare-period/{proof_id}")
async def compare_plans_vs_proof(
    proof_id: int,
    db: AsyncSession = Depends(get_db)
):
    """
    Compare ALL route plans valid in the proof's period against the proof.
    
    This aggregates multiple plans if they exist for the period,
    weighting each by the number of working days it covers.
    """
    # Get proof with details
    proof_result = await db.execute(
        select(Proof)
        .options(
            selectinload(Proof.route_details),
            selectinload(Proof.linehaul_details),
            selectinload(Proof.carrier),
        )
        .where(Proof.id == proof_id)
    )
    proof = proof_result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    # Get all plans for this carrier and period
    plans = await get_plans_for_period(proof.carrier_id, proof.period, db)
    
    if not plans:
        return {
            'proof': {
                'id': proof.id,
                'period': proof.period,
                'carrierId': proof.carrier_id,
                'carrierName': proof.carrier.name if proof.carrier else None,
            },
            'plans': None,
            'status': 'error',
            'message': f'Žádné plány pro období {proof.period}',
            'differences': [],
            'warnings': [{
                'type': 'no_plans',
                'label': 'Chybí plány',
                'note': f'Pro období {proof.period} neexistují žádné platné plány'
            }]
        }
    
    # Aggregate plans
    aggregated = aggregate_plans_for_period(plans, proof.period)
    
    # Count routes from proof
    proof_data = {
        'dpoRoutesCount': 0,
        'sdRoutesCount': 0,
        'sdSpojenCount': 0,
        'drCount': 0,
        'linehaulCount': len(proof.linehaul_details) if proof.linehaul_details else 0,
    }
    
    if proof.route_details:
        for detail in proof.route_details:
            rt = (detail.route_type or '').upper()
            count = detail.routes_count or detail.count or 0
            
            if 'DR' in rt:
                proof_data['drCount'] += count
            elif 'LH_DPO' in rt or rt == 'DPO':
                proof_data['dpoRoutesCount'] += count
            elif 'SPOJENE' in rt or 'SPOJENÉ' in rt:
                proof_data['sdSpojenCount'] += count
            elif 'LH_SD' in rt or rt == 'SD':
                proof_data['sdRoutesCount'] += count
    
    # Build comparison
    differences = []
    warnings = []
    status = 'ok'
    
    # DPO routes comparison
    dpo_diff = proof_data['dpoRoutesCount'] - aggregated['dpoRoutesCount']
    if abs(dpo_diff) > 0:
        differences.append({
            'type': 'dpo_routes',
            'label': 'DPO trasy',
            'planned': aggregated['dpoRoutesCount'],
            'actual': proof_data['dpoRoutesCount'],
            'diff': dpo_diff,
            'note': f"{'Více' if dpo_diff > 0 else 'Méně'} tras než plán ({aggregated['totalWorkingDays']} pracovních dnů)"
        })
        status = 'warning'
    
    # SD routes comparison
    sd_diff = proof_data['sdRoutesCount'] - aggregated['sdRoutesCount']
    if abs(sd_diff) > 0:
        differences.append({
            'type': 'sd_routes',
            'label': 'SD trasy',
            'planned': aggregated['sdRoutesCount'],
            'actual': proof_data['sdRoutesCount'],
            'diff': sd_diff,
            'note': f"{'Více' if sd_diff > 0 else 'Méně'} tras než plán"
        })
        status = 'warning'
    
    # Linehaul comparison
    expected_lh = aggregated['dpoLinehaulCount'] + aggregated['sdLinehaulCount']
    lh_diff = proof_data['linehaulCount'] - expected_lh
    if abs(lh_diff) > 0:
        differences.append({
            'type': 'linehaul',
            'label': 'Linehauly',
            'planned': expected_lh,
            'actual': proof_data['linehaulCount'],
            'diff': lh_diff,
            'note': 'Rozdíl v počtu linehaulů'
        })
    
    # Warnings
    if proof_data['sdSpojenCount'] > 0:
        warnings.append({
            'type': 'merged_routes',
            'label': 'Spojené SD trasy',
            'count': proof_data['sdSpojenCount'],
            'note': f"{proof_data['sdSpojenCount']} spojených SD tras"
        })
    
    if proof_data['drCount'] > 0:
        warnings.append({
            'type': 'direct_routes',
            'label': 'Přímé rozvozy (DR)',
            'count': proof_data['drCount'],
            'note': f"{proof_data['drCount']} přímých rozvozů bez DEPA"
        })
    
    if aggregated['missingDays'] > 0:
        warnings.append({
            'type': 'missing_coverage',
            'label': 'Nepokryté dny',
            'count': aggregated['missingDays'],
            'dates': aggregated['missingDates'][:5],  # First 5
            'note': f"{aggregated['missingDays']} pracovních dnů bez platného plánu"
        })
        status = 'warning'
    
    return {
        'proof': {
            'id': proof.id,
            'period': proof.period,
            'carrierId': proof.carrier_id,
            'carrierName': proof.carrier.name if proof.carrier else None,
            'dpoRoutesCount': proof_data['dpoRoutesCount'],
            'sdRoutesCount': proof_data['sdRoutesCount'],
            'sdSpojenCount': proof_data['sdSpojenCount'],
            'drCount': proof_data['drCount'],
            'linehaulCount': proof_data['linehaulCount'],
        },
        'plans': aggregated,
        'status': status,
        'differences': differences,
        'warnings': warnings,
    }


@router.get("/{plan_id}")
async def get_route_plan(plan_id: int, db: AsyncSession = Depends(get_db)):
    """Get single route plan with all routes"""
    plan = await get_route_plan_by_id(plan_id, db)
    
    if not plan:
        raise HTTPException(status_code=404, detail="Route plan not found")
    
    return {
        'id': plan.id,
        'carrierId': plan.carrier_id,
        'carrierName': plan.carrier.name if plan.carrier else None,
        'validFrom': plan.valid_from.isoformat() if plan.valid_from else None,
        'validTo': plan.valid_to.isoformat() if plan.valid_to else None,
        'fileName': plan.file_name,
        'planType': plan.plan_type,
        'totalRoutes': plan.total_routes,
        'dpoRoutesCount': plan.dpo_routes_count,
        'sdRoutesCount': plan.sd_routes_count,
        'dpoLinehaulCount': plan.dpo_linehaul_count,
        'sdLinehaulCount': plan.sd_linehaul_count,
        'totalDistanceKm': float(plan.total_distance_km) if plan.total_distance_km else 0,
        'totalStops': plan.total_stops,
        'routes': [
            {
                'id': r.id,
                'routeName': r.route_name,
                'routeLetter': r.route_letter,
                'routeType': r.route_type,
                'deliveryType': r.delivery_type,
                'startLocation': r.start_location,
                'stopsCount': r.stops_count,
                'maxCapacity': r.max_capacity,
                'startTime': r.start_time,
                'endTime': r.end_time,
                'workTime': r.work_time,
                'distanceKm': float(r.distance_km) if r.distance_km else 0,
            }
            for r in plan.routes
        ]
    }


@router.get("/{plan_id}/compare/{proof_id}")
async def compare_plan_vs_proof(
    plan_id: int, 
    proof_id: int, 
    db: AsyncSession = Depends(get_db)
):
    """Compare single route plan against proof (legacy endpoint)"""
    # Get plan
    plan = await get_route_plan_by_id(plan_id, db)
    if not plan:
        raise HTTPException(status_code=404, detail="Route plan not found")
    
    # Get proof with details
    proof_result = await db.execute(
        select(Proof)
        .options(
            selectinload(Proof.route_details),
            selectinload(Proof.linehaul_details),
        )
        .where(Proof.id == proof_id)
    )
    proof = proof_result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    # Build comparison
    comparison = {
        'plan': {
            'id': plan.id,
            'planType': plan.plan_type,
            'validFrom': plan.valid_from.isoformat() if plan.valid_from else None,
            'validTo': plan.valid_to.isoformat() if plan.valid_to else None,
            'dpoRoutesCount': plan.dpo_routes_count,
            'sdRoutesCount': plan.sd_routes_count,
            'dpoLinehaulCount': plan.dpo_linehaul_count,
            'sdLinehaulCount': plan.sd_linehaul_count,
        },
        'proof': {
            'id': proof.id,
            'period': proof.period,
            'dpoRoutesCount': 0,
            'sdRoutesCount': 0,
            'sdSpojenCount': 0,
            'linehaulCount': len(proof.linehaul_details) if proof.linehaul_details else 0,
        },
        'differences': [],
        'warnings': [],
        'status': 'ok'
    }
    
    # Count routes by type from proof
    if proof.route_details:
        for detail in proof.route_details:
            rt = (detail.route_type or '').upper()
            count = detail.routes_count or detail.count or 0
            
            if 'LH_DPO' in rt or rt == 'DPO':
                comparison['proof']['dpoRoutesCount'] += count
            elif 'SPOJENE' in rt or 'SPOJENÉ' in rt:
                comparison['proof']['sdSpojenCount'] += count
            elif 'LH_SD' in rt or rt == 'SD':
                comparison['proof']['sdRoutesCount'] += count
    
    # Adjust comparison based on plan type
    if plan.plan_type == 'DPO':
        # Only compare DPO routes
        dpo_diff = comparison['proof']['dpoRoutesCount'] - comparison['plan']['dpoRoutesCount']
        if dpo_diff != 0:
            comparison['differences'].append({
                'type': 'dpo_routes',
                'label': 'DPO trasy',
                'planned': comparison['plan']['dpoRoutesCount'],
                'actual': comparison['proof']['dpoRoutesCount'],
                'diff': dpo_diff,
                'note': f"{'Více' if dpo_diff > 0 else 'Méně'} tras než plán"
            })
            comparison['status'] = 'warning'
    elif plan.plan_type == 'SD':
        # Only compare SD routes
        sd_diff = comparison['proof']['sdRoutesCount'] - comparison['plan']['sdRoutesCount']
        if sd_diff != 0:
            comparison['differences'].append({
                'type': 'sd_routes',
                'label': 'SD trasy',
                'planned': comparison['plan']['sdRoutesCount'],
                'actual': comparison['proof']['sdRoutesCount'],
                'diff': sd_diff,
                'note': f"{'Více' if sd_diff > 0 else 'Méně'} tras než plán"
            })
            comparison['status'] = 'warning'
    else:
        # BOTH - compare both DPO and SD
        dpo_diff = comparison['proof']['dpoRoutesCount'] - comparison['plan']['dpoRoutesCount']
        sd_diff = comparison['proof']['sdRoutesCount'] - comparison['plan']['sdRoutesCount']
        
        if dpo_diff != 0:
            comparison['differences'].append({
                'type': 'dpo_routes',
                'label': 'DPO trasy',
                'planned': comparison['plan']['dpoRoutesCount'],
                'actual': comparison['proof']['dpoRoutesCount'],
                'diff': dpo_diff,
                'note': f"{'Více' if dpo_diff > 0 else 'Méně'} tras než plán"
            })
            comparison['status'] = 'warning'
        
        if sd_diff != 0:
            comparison['differences'].append({
                'type': 'sd_routes',
                'label': 'SD trasy',
                'planned': comparison['plan']['sdRoutesCount'],
                'actual': comparison['proof']['sdRoutesCount'],
                'diff': sd_diff,
                'note': f"{'Více' if sd_diff > 0 else 'Méně'} tras než plán"
            })
            comparison['status'] = 'warning'
    
    if comparison['proof']['sdSpojenCount'] > 0:
        comparison['warnings'].append({
            'type': 'merged_routes',
            'label': 'Spojené trasy',
            'count': comparison['proof']['sdSpojenCount'],
            'note': f"{comparison['proof']['sdSpojenCount']} spojených SD tras"
        })
    
    # Linehaul comparison
    expected_lh = comparison['plan']['dpoLinehaulCount'] + comparison['plan']['sdLinehaulCount']
    if comparison['proof']['linehaulCount'] != expected_lh:
        comparison['differences'].append({
            'type': 'linehaul',
            'label': 'Linehauly',
            'planned': expected_lh,
            'actual': comparison['proof']['linehaulCount'],
            'diff': comparison['proof']['linehaulCount'] - expected_lh,
            'note': 'Rozdíl v počtu linehaulů'
        })
    
    return comparison


@router.get("/daily-breakdown/{proof_id}")
async def get_daily_breakdown(
    proof_id: int,
    db: AsyncSession = Depends(get_db)
):
    """
    Get day-by-day comparison of planned vs actual routes.
    
    Returns each day with:
    - Planned routes from active plan(s)
    - Actual routes and km from proof daily data (total + per depot)
    - Difference
    """
    # Get proof with daily details
    proof_result = await db.execute(
        select(Proof)
        .options(
            selectinload(Proof.carrier),
            selectinload(Proof.daily_details),
        )
        .where(Proof.id == proof_id)
    )
    proof = proof_result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    # Parse period
    try:
        month, year = proof.period.split('/')
        month, year = int(month), int(year)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid period format")
    
    # Get first and last day of month
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, calendar.monthrange(year, month)[1])
    
    # Get all plans that might overlap with this period
    plans = await get_plans_for_period(proof.carrier_id, proof.period, db)
    
    # Build a map of proof daily data by date
    proof_daily_map = {}
    if proof.daily_details:
        for detail in proof.daily_details:
            date_key = detail.date.strftime('%Y-%m-%d')
            proof_daily_map[date_key] = {
                # Počty - celkem
                'drDpo': detail.dr_dpo_count,
                'lhDpo': detail.lh_dpo_count,
                'drSd': detail.dr_sd_count,
                'lhSd': detail.lh_sd_count,
                'totalDpo': detail.dr_dpo_count + detail.lh_dpo_count,
                'totalSd': detail.dr_sd_count + detail.lh_sd_count,
                # Počty - Vratimov
                'vratimovDrDpo': detail.vratimov_dr_dpo,
                'vratimovLhDpo': detail.vratimov_lh_dpo,
                'vratimovDrSd': detail.vratimov_dr_sd,
                'vratimovLhSd': detail.vratimov_lh_sd,
                'vratimovTotal': detail.vratimov_dr_dpo + detail.vratimov_lh_dpo + detail.vratimov_dr_sd + detail.vratimov_lh_sd,
                # Počty - Nový Bydžov
                'bydzovDrDpo': detail.bydzov_dr_dpo,
                'bydzovLhDpo': detail.bydzov_lh_dpo,
                'bydzovDrSd': detail.bydzov_dr_sd,
                'bydzovLhSd': detail.bydzov_lh_sd,
                'bydzovTotal': detail.bydzov_dr_dpo + detail.bydzov_lh_dpo + detail.bydzov_dr_sd + detail.bydzov_lh_sd,
                # Kilometry - celkem
                'drDpoKm': float(detail.dr_dpo_km or 0),
                'lhDpoKm': float(detail.lh_dpo_km or 0),
                'drSdKm': float(detail.dr_sd_km or 0),
                'lhSdKm': float(detail.lh_sd_km or 0),
                'totalDpoKm': float((detail.dr_dpo_km or 0) + (detail.lh_dpo_km or 0)),
                'totalSdKm': float((detail.dr_sd_km or 0) + (detail.lh_sd_km or 0)),
                'totalKm': float((detail.dr_dpo_km or 0) + (detail.lh_dpo_km or 0) + 
                                (detail.dr_sd_km or 0) + (detail.lh_sd_km or 0)),
                # Kilometry - Vratimov
                'vratimovKm': float((detail.vratimov_dr_dpo_km or 0) + (detail.vratimov_lh_dpo_km or 0) +
                                   (detail.vratimov_dr_sd_km or 0) + (detail.vratimov_lh_sd_km or 0)),
                # Kilometry - Nový Bydžov
                'bydzovKm': float((detail.bydzov_dr_dpo_km or 0) + (detail.bydzov_lh_dpo_km or 0) +
                                 (detail.bydzov_dr_sd_km or 0) + (detail.bydzov_lh_sd_km or 0)),
            }
    
    # Build daily breakdown
    days = []
    current = first_day
    
    totals_planned = {
        'totalDays': 0,
        'dpoRoutes': 0,
        'sdRoutes': 0,
    }
    
    totals_actual = {
        'dpoRoutes': 0,
        'sdRoutes': 0,
        'totalKm': 0,
        'vratimovRoutes': 0,
        'bydzovRoutes': 0,
        'vratimovKm': 0,
        'bydzovKm': 0,
    }
    
    while current <= last_day:
        date_key = current.strftime('%Y-%m-%d')
        
        # Find active plan(s) for this day
        active_plans = []
        planned_dpo = 0
        planned_sd = 0
        
        for plan in plans:
            plan_start = plan.valid_from
            plan_end = plan.valid_to or last_day
            
            if plan_start <= current <= plan_end:
                active_plans.append({
                    'id': plan.id,
                    'planType': plan.plan_type,
                    'fileName': plan.file_name,
                })
                
                # Add routes based on plan type
                if plan.plan_type in ('BOTH', 'DPO'):
                    planned_dpo += plan.dpo_routes_count or 0
                if plan.plan_type in ('BOTH', 'SD'):
                    planned_sd += plan.sd_routes_count or 0
        
        # Get actual from proof
        proof_day = proof_daily_map.get(date_key, {})
        actual_dpo = proof_day.get('totalDpo', 0)
        actual_sd = proof_day.get('totalSd', 0)
        actual_km = proof_day.get('totalKm', 0)
        
        # Calculate differences
        diff_dpo = actual_dpo - planned_dpo
        diff_sd = actual_sd - planned_sd
        
        day_data = {
            'date': date_key,
            'dayOfWeek': ['Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne'][current.weekday()],
            'dayNumber': current.day,
            'plans': active_plans,
            # Planned
            'plannedDpo': planned_dpo,
            'plannedSd': planned_sd,
            'plannedTotal': planned_dpo + planned_sd,
            # Actual from proof - počty celkem
            'actualDpo': actual_dpo,
            'actualSd': actual_sd,
            'actualTotal': actual_dpo + actual_sd,
            # Actual from proof - detail počtů
            'actualDrDpo': proof_day.get('drDpo', 0),
            'actualLhDpo': proof_day.get('lhDpo', 0),
            'actualDrSd': proof_day.get('drSd', 0),
            'actualLhSd': proof_day.get('lhSd', 0),
            # Actual - Vratimov
            'vratimovDpo': proof_day.get('vratimovDrDpo', 0) + proof_day.get('vratimovLhDpo', 0),
            'vratimovSd': proof_day.get('vratimovDrSd', 0) + proof_day.get('vratimovLhSd', 0),
            'vratimovTotal': proof_day.get('vratimovTotal', 0),
            'vratimovKm': proof_day.get('vratimovKm', 0),
            # Actual - Nový Bydžov
            'bydzovDpo': proof_day.get('bydzovDrDpo', 0) + proof_day.get('bydzovLhDpo', 0),
            'bydzovSd': proof_day.get('bydzovDrSd', 0) + proof_day.get('bydzovLhSd', 0),
            'bydzovTotal': proof_day.get('bydzovTotal', 0),
            'bydzovKm': proof_day.get('bydzovKm', 0),
            # Actual from proof - kilometry
            'actualDpoKm': proof_day.get('totalDpoKm', 0),
            'actualSdKm': proof_day.get('totalSdKm', 0),
            'actualTotalKm': actual_km,
            # Differences
            'diffDpo': diff_dpo,
            'diffSd': diff_sd,
            'diffTotal': (actual_dpo + actual_sd) - (planned_dpo + planned_sd),
            # Status
            'hasData': bool(proof_day),
            'hasPlan': len(active_plans) > 0,
            'isOk': diff_dpo == 0 and diff_sd == 0,
        }
        
        days.append(day_data)
        
        # Update totals
        totals_planned['totalDays'] += 1
        totals_planned['dpoRoutes'] += planned_dpo
        totals_planned['sdRoutes'] += planned_sd
        totals_actual['dpoRoutes'] += actual_dpo
        totals_actual['sdRoutes'] += actual_sd
        totals_actual['totalKm'] += actual_km
        totals_actual['vratimovRoutes'] += proof_day.get('vratimovTotal', 0)
        totals_actual['bydzovRoutes'] += proof_day.get('bydzovTotal', 0)
        totals_actual['vratimovKm'] += proof_day.get('vratimovKm', 0)
        totals_actual['bydzovKm'] += proof_day.get('bydzovKm', 0)
        
        current += timedelta(days=1)
    
    # Count days with issues
    days_with_diff = sum(1 for d in days if not d['isOk'] and d['hasData'])
    days_without_plan = sum(1 for d in days if not d['hasPlan'])
    days_without_data = sum(1 for d in days if not d['hasData'])
    
    return {
        'proof': {
            'id': proof.id,
            'period': proof.period,
            'carrierId': proof.carrier_id,
            'carrierName': proof.carrier.name if proof.carrier else None,
        },
        'month': {
            'year': year,
            'month': month,
            'monthName': ['', 'Leden', 'Únor', 'Březen', 'Duben', 'Květen', 'Červen', 
                         'Červenec', 'Srpen', 'Září', 'Říjen', 'Listopad', 'Prosinec'][month],
            'totalDays': len(days),
        },
        'days': days,
        'totals': {
            'planned': {
                'dpoRoutes': totals_planned['dpoRoutes'],
                'sdRoutes': totals_planned['sdRoutes'],
                'totalRoutes': totals_planned['dpoRoutes'] + totals_planned['sdRoutes'],
            },
            'actual': {
                'dpoRoutes': totals_actual['dpoRoutes'],
                'sdRoutes': totals_actual['sdRoutes'],
                'totalRoutes': totals_actual['dpoRoutes'] + totals_actual['sdRoutes'],
                'totalKm': totals_actual['totalKm'],
                'vratimovRoutes': totals_actual['vratimovRoutes'],
                'bydzovRoutes': totals_actual['bydzovRoutes'],
                'vratimovKm': totals_actual['vratimovKm'],
                'bydzovKm': totals_actual['bydzovKm'],
            },
            'diff': {
                'dpoRoutes': totals_actual['dpoRoutes'] - totals_planned['dpoRoutes'],
                'sdRoutes': totals_actual['sdRoutes'] - totals_planned['sdRoutes'],
                'totalRoutes': (totals_actual['dpoRoutes'] + totals_actual['sdRoutes']) - 
                              (totals_planned['dpoRoutes'] + totals_planned['sdRoutes']),
            }
        },
        'summary': {
            'daysWithDiff': days_with_diff,
            'daysWithoutPlan': days_without_plan,
            'daysWithoutData': days_without_data,
            'status': 'ok' if days_with_diff == 0 else 'warning',
        }
    }


@router.delete("/{plan_id}", status_code=204)
async def delete_route_plan(plan_id: int, db: AsyncSession = Depends(get_db)):
    """Delete route plan"""
    result = await db.execute(select(RoutePlan).where(RoutePlan.id == plan_id))
    plan = result.scalar_one_or_none()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Route plan not found")
    
    carrier_id = plan.carrier_id
    plan_type = plan.plan_type
    await db.delete(plan)
    
    # Update validity of remaining plans
    await update_route_plan_validity(carrier_id, db, plan_type)
    
    await db.commit()
    
    return None  # Explicit return for 204 No Content
