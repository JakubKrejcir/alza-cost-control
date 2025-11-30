"""
Proofs API Router - with XLSX upload and parsing including daily breakdown
"""
from typing import List, Optional
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import openpyxl

from app.database import get_db
from app.models import (
    Proof, Carrier, ProofRouteDetail, ProofLinehaulDetail, ProofDepoDetail, ProofDailyDetail, ProofAnalysis
)
from app.schemas import ProofResponse, ProofDetailResponse, ProofUpdate

router = APIRouter()


def parse_daily_data_from_xlsx(wb: openpyxl.Workbook) -> List[dict]:
    """
    Parse daily breakdown from 'Podkladove tab' sheet.
    
    Structure:
    - TABLE 1 (CNT - počty tras): Rows 2-61
      - Row 3: Dates (every 4th column starting from C)
      - Row 4: Headers (DR DPO cnt, LH DPO cnt, DR SD cnt, LH SD cnt)
      - Rows 5-39: DEPO VRATIMOV
      - Rows 40-60: DEPO NOVÝ BYDŽOV  
      - Row 61: 'Celkový súčet' - daily totals
      
    - TABLE 2 (KM - kilometry): Rows 64-123
      - Row 65: Dates (every 4th column starting from C)
      - Row 66: Headers (DR DPO km, LH DPO km, DR SD km, LH SD km)
      - Rows 67-101: DEPO VRATIMOV
      - Rows 102-122: DEPO NOVÝ BYDŽOV
      - Row 123: 'Celkový súčet' - daily totals
    """
    # Try different sheet names
    sheet_names = ['Podkladove tab', 'Podkladová tab', 'Podkladove', 'Daily']
    sheet = None
    
    for name in sheet_names:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    
    if sheet is None:
        return []  # No daily data sheet found
    
    # Find depot boundaries by looking for markers in column A
    vratimov_cnt_start = None
    bydzov_cnt_start = None
    cnt_summary_row = None
    
    vratimov_km_start = None
    bydzov_km_start = None
    km_summary_row = None
    
    for row in range(1, min(sheet.max_row + 1, 150)):
        cell_a = sheet.cell(row=row, column=1).value
        cell_b = sheet.cell(row=row, column=2).value
        
        # CNT section
        if cell_a and 'VRATIMOV' in str(cell_a).upper() and vratimov_cnt_start is None:
            vratimov_cnt_start = row
        elif cell_a and ('BYDŽOV' in str(cell_a).upper() or 'BYDZOV' in str(cell_a).upper()):
            if bydzov_cnt_start is None and vratimov_cnt_start is not None and row < 65:
                bydzov_cnt_start = row
            elif bydzov_km_start is None and vratimov_km_start is not None:
                bydzov_km_start = row
        
        # Summary rows
        if cell_b and 'celkov' in str(cell_b).lower() and ('súčet' in str(cell_b).lower() or 'součet' in str(cell_b).lower()):
            if cnt_summary_row is None:
                cnt_summary_row = row
            elif km_summary_row is None:
                km_summary_row = row
        
        # KM section starts
        if cell_a and 'VRATIMOV' in str(cell_a).upper() and vratimov_km_start is None and row > 60:
            vratimov_km_start = row
    
    # Default values if not found
    if vratimov_cnt_start is None:
        vratimov_cnt_start = 5
    if bydzov_cnt_start is None:
        bydzov_cnt_start = 40
    if cnt_summary_row is None:
        cnt_summary_row = 61
    if vratimov_km_start is None:
        vratimov_km_start = 67
    if bydzov_km_start is None:
        bydzov_km_start = 102
    if km_summary_row is None:
        km_summary_row = 123
    
    # Calculate end rows
    vratimov_cnt_end = bydzov_cnt_start - 1
    bydzov_cnt_end = cnt_summary_row - 1
    vratimov_km_end = bydzov_km_start - 1
    bydzov_km_end = km_summary_row - 1
    
    # Parse dates and data
    daily_data = {}
    
    col = 3
    while col <= sheet.max_column:
        date_val = sheet.cell(row=3, column=col).value
        
        if date_val is None:
            col += 4
            continue
        
        # Parse date
        if isinstance(date_val, datetime):
            date = date_val
        elif isinstance(date_val, str):
            try:
                date = datetime.strptime(date_val[:10], '%Y-%m-%d')
            except:
                col += 4
                continue
        else:
            col += 4
            continue
        
        date_key = date.strftime('%Y-%m-%d')
        
        # Helper to sum range
        def sum_range(start_row, end_row, col_offset):
            total = 0
            for r in range(start_row, end_row + 1):
                val = sheet.cell(row=r, column=col + col_offset).value
                if val:
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass
            return total
        
        # CNT - Vratimov
        vratimov_dr_dpo = int(sum_range(vratimov_cnt_start, vratimov_cnt_end, 0))
        vratimov_lh_dpo = int(sum_range(vratimov_cnt_start, vratimov_cnt_end, 1))
        vratimov_dr_sd = int(sum_range(vratimov_cnt_start, vratimov_cnt_end, 2))
        vratimov_lh_sd = int(sum_range(vratimov_cnt_start, vratimov_cnt_end, 3))
        
        # CNT - Nový Bydžov
        bydzov_dr_dpo = int(sum_range(bydzov_cnt_start, bydzov_cnt_end, 0))
        bydzov_lh_dpo = int(sum_range(bydzov_cnt_start, bydzov_cnt_end, 1))
        bydzov_dr_sd = int(sum_range(bydzov_cnt_start, bydzov_cnt_end, 2))
        bydzov_lh_sd = int(sum_range(bydzov_cnt_start, bydzov_cnt_end, 3))
        
        # CNT - Total (from summary row for verification, or sum)
        total_dr_dpo = vratimov_dr_dpo + bydzov_dr_dpo
        total_lh_dpo = vratimov_lh_dpo + bydzov_lh_dpo
        total_dr_sd = vratimov_dr_sd + bydzov_dr_sd
        total_lh_sd = vratimov_lh_sd + bydzov_lh_sd
        
        daily_data[date_key] = {
            'date': date,
            # Celkem
            'dr_dpo_count': total_dr_dpo,
            'lh_dpo_count': total_lh_dpo,
            'dr_sd_count': total_dr_sd,
            'lh_sd_count': total_lh_sd,
            # Vratimov
            'vratimov_dr_dpo': vratimov_dr_dpo,
            'vratimov_lh_dpo': vratimov_lh_dpo,
            'vratimov_dr_sd': vratimov_dr_sd,
            'vratimov_lh_sd': vratimov_lh_sd,
            # Nový Bydžov
            'bydzov_dr_dpo': bydzov_dr_dpo,
            'bydzov_lh_dpo': bydzov_lh_dpo,
            'bydzov_dr_sd': bydzov_dr_sd,
            'bydzov_lh_sd': bydzov_lh_sd,
            # KM - inicializace
            'dr_dpo_km': 0, 'lh_dpo_km': 0, 'dr_sd_km': 0, 'lh_sd_km': 0,
            'vratimov_dr_dpo_km': 0, 'vratimov_lh_dpo_km': 0, 'vratimov_dr_sd_km': 0, 'vratimov_lh_sd_km': 0,
            'bydzov_dr_dpo_km': 0, 'bydzov_lh_dpo_km': 0, 'bydzov_dr_sd_km': 0, 'bydzov_lh_sd_km': 0,
        }
        
        col += 4
    
    # Parse KM data
    col = 3
    while col <= sheet.max_column:
        # KM table has dates in row 65
        date_val = sheet.cell(row=65, column=col).value
        
        if date_val is None:
            col += 4
            continue
        
        # Parse date
        if isinstance(date_val, datetime):
            date = date_val
        elif isinstance(date_val, str):
            try:
                date = datetime.strptime(date_val[:10], '%Y-%m-%d')
            except:
                col += 4
                continue
        else:
            col += 4
            continue
        
        date_key = date.strftime('%Y-%m-%d')
        
        if date_key not in daily_data:
            col += 4
            continue
        
        # Helper to sum range for KM
        def sum_range_km(start_row, end_row, col_offset):
            total = 0.0
            for r in range(start_row, end_row + 1):
                val = sheet.cell(row=r, column=col + col_offset).value
                if val:
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass
            return total
        
        # KM - Vratimov
        daily_data[date_key]['vratimov_dr_dpo_km'] = sum_range_km(vratimov_km_start, vratimov_km_end, 0)
        daily_data[date_key]['vratimov_lh_dpo_km'] = sum_range_km(vratimov_km_start, vratimov_km_end, 1)
        daily_data[date_key]['vratimov_dr_sd_km'] = sum_range_km(vratimov_km_start, vratimov_km_end, 2)
        daily_data[date_key]['vratimov_lh_sd_km'] = sum_range_km(vratimov_km_start, vratimov_km_end, 3)
        
        # KM - Nový Bydžov
        daily_data[date_key]['bydzov_dr_dpo_km'] = sum_range_km(bydzov_km_start, bydzov_km_end, 0)
        daily_data[date_key]['bydzov_lh_dpo_km'] = sum_range_km(bydzov_km_start, bydzov_km_end, 1)
        daily_data[date_key]['bydzov_dr_sd_km'] = sum_range_km(bydzov_km_start, bydzov_km_end, 2)
        daily_data[date_key]['bydzov_lh_sd_km'] = sum_range_km(bydzov_km_start, bydzov_km_end, 3)
        
        # KM - Celkem
        daily_data[date_key]['dr_dpo_km'] = daily_data[date_key]['vratimov_dr_dpo_km'] + daily_data[date_key]['bydzov_dr_dpo_km']
        daily_data[date_key]['lh_dpo_km'] = daily_data[date_key]['vratimov_lh_dpo_km'] + daily_data[date_key]['bydzov_lh_dpo_km']
        daily_data[date_key]['dr_sd_km'] = daily_data[date_key]['vratimov_dr_sd_km'] + daily_data[date_key]['bydzov_dr_sd_km']
        daily_data[date_key]['lh_sd_km'] = daily_data[date_key]['vratimov_lh_sd_km'] + daily_data[date_key]['bydzov_lh_sd_km']
        
        col += 4
    
    # Return as sorted list
    return [daily_data[key] for key in sorted(daily_data.keys())]


def parse_proof_from_xlsx(file_content: bytes) -> dict:
    """Parse proof data from XLSX Sumar sheet and daily data from Podkladove tab"""
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    
    if 'Sumar' not in wb.sheetnames:
        raise ValueError('Sheet "Sumar" not found in XLSX')
    
    sheet = wb['Sumar']
    
    result = {
        'totals': {
            'total_fix': None,
            'total_km': None,
            'total_linehaul': None,
            'total_depo': None,
            'total_penalty': None,
            'grand_total': None,
        },
        'route_details': [],
        'linehaul_details': [],
        'depo_details': [],
        'daily_details': [],
    }
    
    def find_row_by_label(label: str) -> Optional[int]:
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=2).value
            if cell_value and label in str(cell_value):
                return row
        return None
    
    def get_value_by_label(label: str) -> Optional[float]:
        row = find_row_by_label(label)
        if row is None:
            return None
        val = sheet.cell(row=row, column=4).value
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    
    # Extract totals
    result['totals']['total_fix'] = get_value_by_label('Cena FIX')
    result['totals']['total_km'] = get_value_by_label('Cena KM')
    result['totals']['total_linehaul'] = get_value_by_label('Linehaul')
    result['totals']['total_depo'] = get_value_by_label('DEPO')
    result['totals']['total_penalty'] = get_value_by_label('Pokuty')
    result['totals']['grand_total'] = get_value_by_label('Celková částka')
    
    if result['totals']['grand_total'] is None:
        result['totals']['grand_total'] = get_value_by_label('Celkem')
    
    # Extract route details
    route_types = ['DR', 'LH_DPO', 'LH_SD', 'LH_SD_SPOJENE', 'LH SD spojené']
    for rt in route_types:
        row = find_row_by_label(rt)
        if row:
            count = sheet.cell(row=row, column=3).value
            rate = sheet.cell(row=row, column=4).value
            amount = sheet.cell(row=row, column=5).value
            
            if count or amount:
                result['route_details'].append({
                    'route_type': rt.replace(' ', '_').upper(),
                    'count': int(count) if count else 0,
                    'rate': float(rate) if rate else 0,
                    'amount': float(amount) if amount else 0,
                })
    
    # Extract linehaul details
    linehaul_labels = ['CZLC4', 'CZTC1', 'LH-LH', 'Kamion', 'Sólo']
    for label in linehaul_labels:
        row = find_row_by_label(label)
        if row:
            desc = sheet.cell(row=row, column=2).value or label
            days = sheet.cell(row=row, column=3).value
            rate = sheet.cell(row=row, column=4).value
            total = sheet.cell(row=row, column=5).value
            
            if total:
                result['linehaul_details'].append({
                    'description': str(desc),
                    'days': int(days) if days else 0,
                    'rate': float(rate) if rate else 0,
                    'total': float(total) if total else 0,
                })
    
    # Extract depo details
    depo_labels = ['Vratimov', 'Nový Bydžov', 'NB']
    for label in depo_labels:
        row = find_row_by_label(label)
        if row:
            name = sheet.cell(row=row, column=2).value or label
            days = sheet.cell(row=row, column=3).value
            rate = sheet.cell(row=row, column=4).value
            amount = sheet.cell(row=row, column=5).value
            
            if amount:
                rate_type = 'monthly' if 'Bydžov' in str(name) or 'NB' in str(name) else 'daily'
                result['depo_details'].append({
                    'depo_name': str(name),
                    'rate_type': rate_type,
                    'days': int(days) if days else 0,
                    'rate': float(rate) if rate else 0,
                    'amount': float(amount) if amount else 0,
                })
    
    # Parse daily data from Podkladove tab
    result['daily_details'] = parse_daily_data_from_xlsx(wb)
    
    return result


@router.get("", response_model=List[ProofResponse])
async def get_proofs(
    carrier_id: Optional[int] = Query(None),
    period: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Get all proofs with filters"""
    query = select(Proof).options(
        selectinload(Proof.carrier),
        selectinload(Proof.depot)
    )
    
    filters = []
    if carrier_id:
        filters.append(Proof.carrier_id == carrier_id)
    if period:
        filters.append(Proof.period == period)
    if status:
        filters.append(Proof.status == status)
    
    if filters:
        query = query.where(and_(*filters))
    
    query = query.order_by(Proof.period_date.desc())
    
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/upload", response_model=ProofDetailResponse, status_code=201)
async def upload_proof(
    file: UploadFile = File(...),
    carrier_id: int = Form(...),
    period: str = Form(...),
    db: AsyncSession = Depends(get_db)
):
    """Upload and parse proof XLSX"""
    # Validate carrier
    carrier_result = await db.execute(
        select(Carrier).where(Carrier.id == carrier_id)
    )
    if not carrier_result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Carrier not found")
    
    content = await file.read()
    
    try:
        proof_data = parse_proof_from_xlsx(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {str(e)}")
    
    # Parse period
    try:
        month, year = period.split('/')
        period_date = datetime(int(year), int(month), 1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid period format. Use MM/YYYY")
    
    # Check for existing proof and delete it
    existing_result = await db.execute(
        select(Proof).where(
            and_(
                Proof.carrier_id == carrier_id,
                Proof.period == period
            )
        )
    )
    existing = existing_result.scalar_one_or_none()
    if existing:
        await db.delete(existing)
        await db.flush()
    
    # Create new proof
    proof = Proof(
        carrier_id=carrier_id,
        period=period,
        period_date=period_date,
        file_name=file.filename,
        status='uploaded',
        total_fix=proof_data['totals'].get('total_fix'),
        total_km=proof_data['totals'].get('total_km'),
        total_linehaul=proof_data['totals'].get('total_linehaul'),
        total_depo=proof_data['totals'].get('total_depo'),
        total_penalty=proof_data['totals'].get('total_penalty'),
        grand_total=proof_data['totals'].get('grand_total'),
    )
    db.add(proof)
    await db.flush()
    
    # Add route details
    for rd in proof_data['route_details']:
        detail = ProofRouteDetail(
            proof_id=proof.id,
            route_type=rd['route_type'],
            routes_count=rd['count'],
            count=rd['count'],
            rate=Decimal(str(rd['rate'])),
            amount=Decimal(str(rd['amount'])),
        )
        db.add(detail)
    
    # Add linehaul details
    for ld in proof_data['linehaul_details']:
        detail = ProofLinehaulDetail(
            proof_id=proof.id,
            description=ld['description'],
            days=ld['days'],
            rate=Decimal(str(ld['rate'])),
            total=Decimal(str(ld['total'])),
        )
        db.add(detail)
    
    # Add depo details
    for dd in proof_data['depo_details']:
        detail = ProofDepoDetail(
            proof_id=proof.id,
            depo_name=dd['depo_name'],
            rate_type=dd['rate_type'],
            days=dd['days'],
            rate=Decimal(str(dd['rate'])),
            amount=Decimal(str(dd['amount'])),
        )
        db.add(detail)
    
    # Add daily details
    for daily in proof_data['daily_details']:
        detail = ProofDailyDetail(
            proof_id=proof.id,
            date=daily['date'],
            # Celkem
            dr_dpo_count=daily['dr_dpo_count'],
            lh_dpo_count=daily['lh_dpo_count'],
            dr_sd_count=daily['dr_sd_count'],
            lh_sd_count=daily['lh_sd_count'],
            # Vratimov
            vratimov_dr_dpo=daily.get('vratimov_dr_dpo', 0),
            vratimov_lh_dpo=daily.get('vratimov_lh_dpo', 0),
            vratimov_dr_sd=daily.get('vratimov_dr_sd', 0),
            vratimov_lh_sd=daily.get('vratimov_lh_sd', 0),
            # Nový Bydžov
            bydzov_dr_dpo=daily.get('bydzov_dr_dpo', 0),
            bydzov_lh_dpo=daily.get('bydzov_lh_dpo', 0),
            bydzov_dr_sd=daily.get('bydzov_dr_sd', 0),
            bydzov_lh_sd=daily.get('bydzov_lh_sd', 0),
            # KM - Celkem
            dr_dpo_km=Decimal(str(daily['dr_dpo_km'])) if daily['dr_dpo_km'] else Decimal('0'),
            lh_dpo_km=Decimal(str(daily['lh_dpo_km'])) if daily['lh_dpo_km'] else Decimal('0'),
            dr_sd_km=Decimal(str(daily['dr_sd_km'])) if daily['dr_sd_km'] else Decimal('0'),
            lh_sd_km=Decimal(str(daily['lh_sd_km'])) if daily['lh_sd_km'] else Decimal('0'),
            # KM - Vratimov
            vratimov_dr_dpo_km=Decimal(str(daily.get('vratimov_dr_dpo_km', 0))),
            vratimov_lh_dpo_km=Decimal(str(daily.get('vratimov_lh_dpo_km', 0))),
            vratimov_dr_sd_km=Decimal(str(daily.get('vratimov_dr_sd_km', 0))),
            vratimov_lh_sd_km=Decimal(str(daily.get('vratimov_lh_sd_km', 0))),
            # KM - Nový Bydžov
            bydzov_dr_dpo_km=Decimal(str(daily.get('bydzov_dr_dpo_km', 0))),
            bydzov_lh_dpo_km=Decimal(str(daily.get('bydzov_lh_dpo_km', 0))),
            bydzov_dr_sd_km=Decimal(str(daily.get('bydzov_dr_sd_km', 0))),
            bydzov_lh_sd_km=Decimal(str(daily.get('bydzov_lh_sd_km', 0))),
        )
        db.add(detail)
    
    await db.commit()
    
    # Reload with relationships
    result = await db.execute(
        select(Proof)
        .options(
            selectinload(Proof.carrier),
            selectinload(Proof.route_details),
            selectinload(Proof.linehaul_details),
            selectinload(Proof.depo_details),
            selectinload(Proof.daily_details),
            selectinload(Proof.invoices),
            selectinload(Proof.analyses),
        )
        .where(Proof.id == proof.id)
    )
    return result.scalar_one()


@router.get("/{proof_id}", response_model=ProofDetailResponse)
async def get_proof(proof_id: int, db: AsyncSession = Depends(get_db)):
    """Get proof with all details"""
    result = await db.execute(
        select(Proof)
        .options(
            selectinload(Proof.carrier),
            selectinload(Proof.depot),
            selectinload(Proof.route_details),
            selectinload(Proof.linehaul_details),
            selectinload(Proof.depo_details),
            selectinload(Proof.daily_details),
            selectinload(Proof.invoices),
        )
        .where(Proof.id == proof_id)
    )
    proof = result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    return proof


@router.get("/{proof_id}/daily")
async def get_proof_daily_details(proof_id: int, db: AsyncSession = Depends(get_db)):
    """Get daily breakdown for a proof (counts and km)"""
    result = await db.execute(
        select(Proof)
        .options(selectinload(Proof.daily_details))
        .where(Proof.id == proof_id)
    )
    proof = result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    daily_data = []
    for detail in sorted(proof.daily_details, key=lambda x: x.date):
        daily_data.append({
            'date': detail.date.strftime('%Y-%m-%d'),
            # Počty
            'drDpoCount': detail.dr_dpo_count,
            'lhDpoCount': detail.lh_dpo_count,
            'drSdCount': detail.dr_sd_count,
            'lhSdCount': detail.lh_sd_count,
            'totalDpo': detail.dr_dpo_count + detail.lh_dpo_count,
            'totalSd': detail.dr_sd_count + detail.lh_sd_count,
            'totalRoutes': detail.dr_dpo_count + detail.lh_dpo_count + detail.dr_sd_count + detail.lh_sd_count,
            # Kilometry
            'drDpoKm': float(detail.dr_dpo_km or 0),
            'lhDpoKm': float(detail.lh_dpo_km or 0),
            'drSdKm': float(detail.dr_sd_km or 0),
            'lhSdKm': float(detail.lh_sd_km or 0),
            'totalDpoKm': float((detail.dr_dpo_km or 0) + (detail.lh_dpo_km or 0)),
            'totalSdKm': float((detail.dr_sd_km or 0) + (detail.lh_sd_km or 0)),
            'totalKm': float((detail.dr_dpo_km or 0) + (detail.lh_dpo_km or 0) + (detail.dr_sd_km or 0) + (detail.lh_sd_km or 0)),
        })
    
    return {
        'proofId': proof.id,
        'period': proof.period,
        'days': daily_data,
        'totals': {
            # Počty
            'drDpo': sum(d['drDpoCount'] for d in daily_data),
            'lhDpo': sum(d['lhDpoCount'] for d in daily_data),
            'drSd': sum(d['drSdCount'] for d in daily_data),
            'lhSd': sum(d['lhSdCount'] for d in daily_data),
            'totalDpo': sum(d['totalDpo'] for d in daily_data),
            'totalSd': sum(d['totalSd'] for d in daily_data),
            'totalRoutes': sum(d['totalRoutes'] for d in daily_data),
            # Kilometry
            'drDpoKm': sum(d['drDpoKm'] for d in daily_data),
            'lhDpoKm': sum(d['lhDpoKm'] for d in daily_data),
            'drSdKm': sum(d['drSdKm'] for d in daily_data),
            'lhSdKm': sum(d['lhSdKm'] for d in daily_data),
            'totalDpoKm': sum(d['totalDpoKm'] for d in daily_data),
            'totalSdKm': sum(d['totalSdKm'] for d in daily_data),
            'totalKm': sum(d['totalKm'] for d in daily_data),
        }
    }


@router.delete("/{proof_id}", status_code=204)
async def delete_proof(proof_id: int, db: AsyncSession = Depends(get_db)):
    """Delete a proof"""
    result = await db.execute(
        select(Proof).where(Proof.id == proof_id)
    )
    proof = result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    await db.delete(proof)
    await db.commit()
    return None


@router.patch("/{proof_id}", response_model=ProofResponse)
async def update_proof(
    proof_id: int,
    update: ProofUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update proof status"""
    result = await db.execute(
        select(Proof).where(Proof.id == proof_id)
    )
    proof = result.scalar_one_or_none()
    if not proof:
        raise HTTPException(status_code=404, detail="Proof not found")
    
    if update.status:
        proof.status = update.status
    
    await db.commit()
    await db.refresh(proof)
    return proof
